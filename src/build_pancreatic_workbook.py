#!/usr/bin/env python3
from __future__ import annotations

import re
import unicodedata
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
import pdfplumber
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

logging.getLogger("pypdf").setLevel(logging.ERROR)
logging.getLogger("pdfminer").setLevel(logging.ERROR)
logging.getLogger("pdfminer.pdffont").setLevel(logging.ERROR)

ROOT = Path('/Users/alfred/Desktop/paper')
OUT_XLSX = ROOT / '胰腺癌_证据抽取_任务123_最终版.xlsx'
QUALITY_DIR = ROOT / 'downloads' / 'quality_run_20260209_fast'

METHOD_ORDER = {
    '化疗': 1,
    '放化疗': 2,
    '靶向治疗': 3,
    '免疫治疗': 4,
    '创新疗法': 5,
}

LAYER_ORDER = {'core': 0, 'extended': 1}

PDF_MAP = {
    'ACCORD 11': ROOT / 'ACCORD 11 NEJM2011.pdf',
    'MPACT': ROOT / 'MPACT.pdf',
    'NEOPAN': ROOT / '2025-JCO-NEOPAN：FOLFIRINOX vs GEM.pdf',
    'NAPOLI-3': ROOT / 'NAPOLI 3 Lancet-2023.pdf',
    'JCOG1611 (GENERATE)': ROOT / 'GENERATE.pdf',
    'GEST': ROOT / 'GEST QOL -ESMO open 2017.pdf',
    'GEST++(IPD)': ROOT / 'GEST-LAPC.pdf',
    'AVENGER 500': ROOT / 'AVENGER 500 Study _ JCO 2024.pdf',
    'NOTABLE': ROOT / 'NOTABLE-JCO-2023.pdf',
    'HALO-301': ROOT / 'HALO-3 JCO-2020.pdf',
    'PA.3': ROOT / 'PA.3 G±Erlotinib JCO-2007.pdf',
    'KG4/2015': ROOT / 'KG4-2015_gv1001 BJC-2024.pdf',
    'CONKO-007': ROOT / 'CONKO-007.pdf',
    'LAP-07': ROOT / 'LAP07_  JAMA-2016.pdf',
}

STUDIES: Dict[str, Dict[str, str]] = {
    'PANOVA-3': {
        '方式': '创新疗法', '类别': '5', '研究时间': '2018-2023', '随访时间/月': '至2024年10月',
        '研究人群': 'LAPC', 'n': '571', '治疗方案': 'TTFields+GnP vs GnP', '发表时间': '2025, JCO',
        '研究层级(core/extended)': 'core', '是否新增文献': '否',
    },
    'ACCORD 11': {
        '方式': '化疗', '类别': '1', '研究时间': '2005-2009', '随访时间/月': '26.6',
        '研究人群': 'MPC (ECOG 0-1)', 'n': '342', '治疗方案': 'FOLFIRINOX vs Gem', '发表时间': '2011, NEJM',
        '研究层级(core/extended)': 'core', '是否新增文献': '否',
    },
    'MPACT': {
        '方式': '化疗', '类别': '1', '研究时间': '2009-2012', '随访时间/月': '9.1',
        '研究人群': 'MPC (KPS≥70)', 'n': '861', '治疗方案': 'GnP vs Gem', '发表时间': '2013, NEJM',
        '研究层级(core/extended)': 'core', '是否新增文献': '否',
    },
    'NEOPAN': {
        '方式': '化疗', '类别': '1', '研究时间': '2015-2022', '随访时间/月': '59.6',
        '研究人群': 'LAPC', 'n': '171', '治疗方案': 'mFOLFIRINOX vs Gem', '发表时间': '2025, JCO',
        '研究层级(core/extended)': 'core', '是否新增文献': '否',
    },
    'NAPOLI-3': {
        '方式': '化疗', '类别': '1', '研究时间': '2020-2021', '随访时间/月': '16.1',
        '研究人群': 'MPC', 'n': '770', '治疗方案': 'NALIRIFOX vs GnP', '发表时间': '2023, Lancet',
        '研究层级(core/extended)': 'core', '是否新增文献': '否',
    },
    'JCOG1611 (GENERATE)': {
        '方式': '化疗', '类别': '1', '研究时间': '2019-2023', '随访时间/月': '10.1',
        '研究人群': 'MPC/rPC', 'n': '527', '治疗方案': 'mFOLFIRINOX vs S-IROX vs GnP', '发表时间': '2025, JCO',
        '研究层级(core/extended)': 'core', '是否新增文献': '否',
    },
    'GEST': {
        '方式': '化疗', '类别': '1', '研究时间': '2007-2009', '随访时间/月': '18.4',
        '研究人群': 'LAPC/MPC', 'n': '834', '治疗方案': 'Gem vs S-1 vs GS', '发表时间': '2013, JCO; 2017, ESMO Open',
        '研究层级(core/extended)': 'core', '是否新增文献': '否',
    },
    'GEST++(IPD)': {
        '方式': '化疗', '类别': '1', '研究时间': '2006-2010', '随访时间/月': '14.3',
        '研究人群': 'LAPC', 'n': '193', '治疗方案': 'Gem vs GS', '发表时间': '2017, BJC',
        '研究层级(core/extended)': 'core', '是否新增文献': '否',
    },
    'AVENGER 500': {
        '方式': '靶向治疗', '类别': '3', '研究时间': '2018-2021', '随访时间/月': '18.7 vs 19.2',
        '研究人群': 'MPC', 'n': '528', '治疗方案': 'mFOLFIRINOX+Devimistat vs FOLFIRINOX', '发表时间': '2024, JCO',
        '研究层级(core/extended)': 'core', '是否新增文献': '否',
    },
    'NOTABLE': {
        '方式': '靶向治疗', '类别': '3', '研究时间': '2015-2021', '随访时间/月': '57.6 vs 16.6',
        '研究人群': 'KRAS野生型 LAPC/MPC', 'n': '82/480', '治疗方案': 'Gem+尼妥珠单抗 vs Gem', '发表时间': '2023, JCO',
        '研究层级(core/extended)': 'core', '是否新增文献': '否',
    },
    'HALO-301': {
        '方式': '靶向治疗', '类别': '3', '研究时间': '2016-2018', '随访时间/月': '至2019年10月',
        '研究人群': 'MPC', 'n': '494', '治疗方案': 'GnP+PEGPH20 vs GnP', '发表时间': '2020, JCO',
        '研究层级(core/extended)': 'core', '是否新增文献': '否',
    },
    'PA.3': {
        '方式': '靶向治疗', '类别': '3', '研究时间': '2001-2003', '随访时间/月': '至2005年6月',
        '研究人群': 'LAPC/MPC', 'n': '569', '治疗方案': 'Gem+Erlotinib vs Gem', '发表时间': '2007, JCO',
        '研究层级(core/extended)': 'core', '是否新增文献': '否',
    },
    'TeloVac': {
        '方式': '免疫治疗', '类别': '4', '研究时间': '2007-2011', '随访时间/月': '6',
        '研究人群': 'LAPC/MPC', 'n': '1572', '治疗方案': 'GemCap vs GemCap序贯CIT vs GemCap+GV1001', '发表时间': '2014, Lancet Oncol',
        '研究层级(core/extended)': 'core', '是否新增文献': '否',
    },
    'KG4/2015': {
        '方式': '免疫治疗', '类别': '4', '研究时间': '2015-2020', '随访时间/月': '7.9',
        '研究人群': 'LAPC/MPC (Eotaxin高表达)', 'n': '148', '治疗方案': 'GemCap+GV1001 vs GemCap', '发表时间': '2024, BJC',
        '研究层级(core/extended)': 'core', '是否新增文献': '否',
    },
    'CONKO-007': {
        '方式': '放化疗', '类别': '2', '研究时间': '2013-2021', '随访时间/月': '76',
        '研究人群': 'LAPC', 'n': '525', '治疗方案': '诱导化疗后 CRT vs ChT', '发表时间': '2025, JCO',
        '研究层级(core/extended)': 'core', '是否新增文献': '否',
    },
    'LAP-07': {
        '方式': '放化疗', '类别': '2', '研究时间': '2008-2011', '随访时间/月': '36.7',
        '研究人群': 'LAPC', 'n': '449', '治疗方案': 'Gem序贯CRT vs Gem', '发表时间': '2016, JAMA',
        '研究层级(core/extended)': 'core', '是否新增文献': '否',
    },
    'ACCORD11_QOL': {
        '方式': '化疗', '类别': '1', '研究时间': '2005-2009', '随访时间/月': 'QoL每2周',
        '研究人群': 'MPC (PS 0-1)', 'n': '342', '治疗方案': 'FOLFIRINOX vs Gem (QoL伴随分析)', '发表时间': '2013, JCO',
        '研究层级(core/extended)': 'extended', '是否新增文献': '是',
    },
    'PA3_QOL': {
        '方式': '靶向治疗', '类别': '3', '研究时间': '2001-2003', '随访时间/月': '基线+8周变化',
        '研究人群': '晚期胰腺癌', 'n': '569', '治疗方案': 'G+E vs G+P (QoL预后分析)', '发表时间': '2016, Pancreatology',
        '研究层级(core/extended)': 'extended', '是否新增文献': '是',
    },
    'MPACT_HRQoL': {
        '方式': '化疗', '类别': '1', '研究时间': '2009-2012', '随访时间/月': '真实世界横断',
        '研究人群': '转移性胰腺癌', 'n': '861(关联CA046)', '治疗方案': 'mPC患者HRQoL评估', '发表时间': '2017, Pancreas',
        '研究层级(core/extended)': 'extended', '是否新增文献': '是',
    },
    'MPACT_NICE_CE': {
        '方式': '化疗', '类别': '1', '研究时间': '基于CA046/NICE复评', '随访时间/月': '',
        '研究人群': 'untreated mPC (技术评估)', 'n': '861(来源CA046)', '治疗方案': 'Nab-Pac+Gem vs Gem/GemCap/FOLFIRINOX', '发表时间': '2018, PharmacoEconomics',
        '研究层级(core/extended)': 'extended', '是否新增文献': '是',
    },
    'APICE_2018': {
        '方式': '化疗', '类别': '1', '研究时间': '4年Markov模型', '随访时间/月': '',
        '研究人群': 'MPC (意大利卫生体系视角)', 'n': '模型研究', '治疗方案': 'Nab-Pac+Gem vs Gem', '发表时间': '2018, Expert Rev Pharmacoecon Outcomes Res',
        '研究层级(core/extended)': 'extended', '是否新增文献': '是',
    },
    'FOLFIRINOX_vs_GNP_CE': {
        '方式': '化疗', '类别': '1', '研究时间': '基于PRODIGE/MPACT', '随访时间/月': '',
        '研究人群': 'MPC (中国成本效果分析)', 'n': '模型研究', '治疗方案': 'FOLFIRINOX vs GEM-N', '发表时间': '2016, Tumori',
        '研究层级(core/extended)': 'extended', '是否新增文献': '是',
    },
    'LAPACT': {
        '方式': '化疗', '类别': '1', '研究时间': '多中心II期', '随访时间/月': '',
        '研究人群': 'unresectable LAPC', 'n': '107', '治疗方案': 'nab-Paclitaxel+Gem induction', '发表时间': '2020, Lancet Gastroenterol Hepatol',
        '研究层级(core/extended)': 'extended', '是否新增文献': '是',
    },
}

TASK1_METRICS = ['mOS/月', 'OS_HR', '1yOS/%', '5yOS/%', 'mPFS/月', '6mPFS/%', '12mPFS/%', 'DFS/%']
TASK2_METRICS = ['ORR/%', 'CR/%', 'DCR/%', 'TTP/月', 'DOR/月', '手术率/%', 'R0切除率/%', 'pCR/%', '局部复发率/%', '远处转移率/%', 'CBR/%', '其他肿瘤控制指标']
TASK3_METRICS = ['≥3级AE/%', '治疗相关死亡/%', 'QOL', 'QALY/QALM', '痛苦评分/症状', 'PRO工具', '关键AE信号']
AUDIT_COLS = ['dimension_id', 'dimension_version', 'definition_source', 'value_source', 'source_tier', 'institution_tier', 'country_group']
MCKINSEY_BLUE = "003A70"

FIELD_GLOSSARY: Dict[str, str] = {
    '方式': '治疗方式大类（化疗/放化疗/靶向/免疫/创新等）',
    '类别': '方式对应的分组编号',
    '临床试验': '试验名称或研究标识',
    '研究时间': '入组或研究执行时间区间',
    '随访时间/月': '中位或关键随访时长（单位：月）',
    '研究人群': '纳入人群定义（如LAPC/MPC/亚组）',
    'n': '样本量',
    '治疗方案': '干预组与对照组方案',
    '发表时间': '论文发表年份与期刊',
    '研究层级(core/extended)': '证据分层：核心/扩展',
    '是否新增文献': '是否为本轮新增纳入文献',
    'mOS/月': '中位总生存期（Overall Survival）',
    'OS_HR': '总生存风险比（Hazard Ratio）',
    '1yOS/%': '1年总生存率',
    '5yOS/%': '5年总生存率',
    'mPFS/月': '中位无进展生存期（Progression-Free Survival）',
    '6mPFS/%': '6个月无进展生存率',
    '12mPFS/%': '12个月无进展生存率',
    'DFS/%': '无病生存相关指标',
    'ORR/%': '客观缓解率（Objective Response Rate）',
    'CR/%': '完全缓解率',
    'DCR/%': '疾病控制率',
    'TTP/月': '进展时间（Time To Progression）',
    'DOR/月': '缓解持续时间（Duration Of Response）',
    '手术率/%': '手术实施比例',
    'R0切除率/%': 'R0切除比例（显微镜下切缘阴性）',
    'pCR/%': '病理完全缓解率',
    '局部复发率/%': '局部复发比例',
    '远处转移率/%': '远处转移比例',
    'CBR/%': '临床获益率（仅任务2使用）',
    '其他肿瘤控制指标': '不在固定列中的肿瘤控制补充指标',
    '≥3级AE/%': '3级及以上不良事件比例',
    '治疗相关死亡/%': '治疗相关死亡比例',
    'QOL': '生活质量结论性指标',
    'QALY/QALM': '质量调整生命年/质量调整生存月',
    '痛苦评分/症状': '疼痛或症状负担相关指标',
    'PRO工具': '患者报告结局量表（如EORTC QLQ-C30/EQ-5D）',
    '关键AE信号': '关键安全性信号摘要',
    '证据ID': '主表单元格对应底表证据ID',
    '证据等级': '证据强度等级（A/B/C）',
    '备注': '补充说明',
    'datum_id': '底表证据唯一ID',
    'endpoint': '证据对应终点名称',
    'value_text': '结构化值文本',
    'population': '证据对应人群',
    'arm_compare': '证据对应组间比较',
    'source_type': '证据来源类型（全文/摘要/截图等）',
    'doi/pmid': '文献DOI或PMID',
    'file_path': '本地文件路径或来源定位',
    'page_no': '页码或可定位片段',
    'quote_original': '原文摘录',
    'evidence_level': '证据等级',
    'task_id': '证据所属任务',
    'trial_id': '证据所属试验',
    'extractor_a_quote': '抽取通道A摘录',
    'extractor_b_quote': '抽取通道B摘录',
    'consistency': '双抽取一致性',
    'adjudication_note': '人工裁决说明',
    '文献': '新增下载记录对应文献标题',
    '来源': '下载来源渠道',
    'download_status': '下载结果状态',
    '本地路径': '本地文件路径',
    '是否已深度提取': '是否已进入深度提取流程',
    '本地存在性': '本地文件是否存在',
    '字段': '错误校验中被比对字段名',
    '截图原值': '用户截图原始值',
    '核验后值': '核验后的当前值',
    '状态': '一致性状态',
    '原文页码': '核验证据页码',
    '原文摘录': '核验证据原文',
    '说明': '校验解释',
    '发现类型': '反直觉发现类型',
    '触发条件': '反直觉发现触发规则',
    '证据摘录': '反直觉发现证据片段',
    'arm_compare': 'AE矩阵组间比较',
    'AE术语': '不良事件术语',
    '分级': 'AE分级',
    '干预组数值': '干预组数值',
    '对照组数值': '对照组数值',
    '单位': '数值单位',
    '是否关键AE': '是否为关键安全性事件',
    'uid': '候选文献唯一ID',
    'title': '文献标题',
    'source': '数据来源平台',
    'year': '发表年份',
    'doi': '数字对象标识符',
    'pmid': 'PubMed ID',
    'pmcid': 'PubMed Central ID',
    'journal': '期刊名称',
    'discipline_profile': '学科场景标签（临床/真实世界/经济学）',
    'source_cred': '来源可信分',
    'journal_cred': '期刊可信分',
    'citation_cred': '引用可信分',
    'design_cred': '研究设计可信分',
    'integrity_cred': '完整性可信分',
    'quality_penalty': '质量惩罚分',
    'quality_penalty_reasons': '惩罚原因',
    'credibility_score': '综合可信度评分',
    'credibility_tier': '可信度分层',
    'quality_gate': '质量门控结果（core/extended/reject）',
    'rejection_reason': '被拒绝原因',
    'journal_tier': '期刊等级（A/B/C/U）',
    'citation_age_years': '发表距今年数',
    'citation_age_adjusted': '年龄校正后引用指标',
    'cited_by_count': '引用次数',
    'preprint_flag': '是否预印本',
    'retracted_flag': '是否撤稿',
    'institution_signal': '机构信号强度',
    'dimension_id': '指标维度ID',
    'dimension_version': '维度版本',
    'definition_source': '维度定义来源',
    'value_source': '具体数据值来源',
    'source_tier': '来源分级（S/A/B/C）',
    'institution_tier': '机构层级（top/high等）',
    'country_group': '国家分组（发达市场/中国顶级中心/其他）',
}

FIELD_PRIORITY = [
    '临床试验', '方式', '类别', '治疗方案', '研究人群', 'n', '研究时间', '随访时间/月', '发表时间',
    '研究层级(core/extended)', '是否新增文献',
    'mOS/月', 'OS_HR', '1yOS/%', '5yOS/%', 'mPFS/月', '6mPFS/%', '12mPFS/%', 'DFS/%',
    'ORR/%', 'CR/%', 'DCR/%', 'TTP/月', 'DOR/月', '手术率/%', 'R0切除率/%', 'pCR/%', '局部复发率/%', '远处转移率/%', 'CBR/%',
    '≥3级AE/%', '治疗相关死亡/%', 'QOL', 'QALY/QALM', '痛苦评分/症状', 'PRO工具', '关键AE信号',
    '证据ID', '证据等级', '备注',
    'datum_id', 'endpoint', 'value_text', 'population', 'arm_compare', 'source_type', 'doi/pmid', 'file_path', 'page_no', 'quote_original', 'evidence_level',
    'dimension_id', 'dimension_version', 'definition_source', 'value_source', 'source_tier', 'institution_tier', 'country_group',
    'quality_gate', 'credibility_score', 'credibility_tier', 'rejection_reason', 'journal_tier', 'cited_by_count',
]
FIELD_PRIORITY_RANK = {name: idx for idx, name in enumerate(FIELD_PRIORITY)}

TASK1_ROWS_RAW = [
    {'临床试验': 'PANOVA-3', 'mOS/月': '16.2 vs 14.2', 'OS_HR': '0.82', '1yOS/%': '68.1 vs 60.2', '5yOS/%': '', 'mPFS/月': '10.6 vs 9.3', '6mPFS/%': '', '12mPFS/%': '43.9 vs 34.1', 'DFS/%': '', '证据ID': 'E001', '证据等级': 'B/C', '备注': '当前以会议摘要与截图为主'},
    {'临床试验': 'ACCORD 11', 'mOS/月': '11.1 vs 6.8', 'OS_HR': '0.57', '1yOS/%': '48.4 vs 20.6', '5yOS/%': '', 'mPFS/月': '6.4 vs 3.3', '6mPFS/%': '52.8 vs 17.2', '12mPFS/%': '12.1 vs 3.5', 'DFS/%': '', '证据ID': 'E002', '证据等级': 'A', '备注': ''},
    {'临床试验': 'MPACT', 'mOS/月': '8.5 vs 6.7', 'OS_HR': '0.72', '1yOS/%': '35 vs 22', '5yOS/%': '', 'mPFS/月': '5.5 vs 3.7', '6mPFS/%': '44 vs 25', '12mPFS/%': '16 vs 9', 'DFS/%': '', '证据ID': 'E003', '证据等级': 'A', '备注': ''},
    {'临床试验': 'NEOPAN', 'mOS/月': '15.7 vs 15.4', 'OS_HR': '1.00', '1yOS/%': '61.2 vs 60.5', '5yOS/%': '', 'mPFS/月': '9.7 vs 7.7', '6mPFS/%': '', '12mPFS/%': '36.5 vs 17.4', 'DFS/%': '', '证据ID': 'E004', '证据等级': 'A', '备注': ''},
    {'临床试验': 'NAPOLI-3', 'mOS/月': '11.1 vs 9.2', 'OS_HR': '0.83', '1yOS/%': '45.6 vs 39.5', '5yOS/%': '', 'mPFS/月': '7.4 vs 5.6', '6mPFS/%': '56.4 vs 43.2', '12mPFS/%': '27.4 vs 13.9', 'DFS/%': '', '证据ID': 'E005', '证据等级': 'A', '备注': ''},
    {'临床试验': 'JCOG1611 (GENERATE)', 'mOS/月': '14.0 vs 13.6 vs 17.0', 'OS_HR': 'mFOLFIRINOX vs GnP:1.29; S-IROX vs GnP:1.29', '1yOS/%': '', '5yOS/%': '', 'mPFS/月': '5.8 vs 6.7 vs 6.7', '6mPFS/%': '', '12mPFS/%': '', 'DFS/%': '', '证据ID': 'E006', '证据等级': 'A', '备注': '三臂研究'},
    {'临床试验': 'GEST', 'mOS/月': '8.8 vs 9.7 vs 10.1', 'OS_HR': '', '1yOS/%': '35.4 vs 38.7 vs 40.7', '5yOS/%': '', 'mPFS/月': '4.1 vs 3.8 vs 5.7', '6mPFS/%': '29.8 vs 26.9 vs 47.9', '12mPFS/%': '9.1 vs 7.2 vs 20.3', 'DFS/%': '', '证据ID': 'E007;E024', '证据等级': 'A/C', '备注': '生存数值含截图复核项'},
    {'临床试验': 'GEST++(IPD)', 'mOS/月': '11.83 vs 16.41', 'OS_HR': '0.708', '1yOS/%': '48.45 vs 67.71', '5yOS/%': '', 'mPFS/月': '5.78 vs 11.76', '6mPFS/%': '48.45 vs 71.42', '12mPFS/%': '17.53 vs 46.73', 'DFS/%': '', '证据ID': 'E008', '证据等级': 'A', '备注': 'LAPC亚组IPD'},
    {'临床试验': 'AVENGER 500', 'mOS/月': '11.10 vs 11.73', 'OS_HR': '0.95', '1yOS/%': '', '5yOS/%': '', 'mPFS/月': '7.82 vs 7.98', '6mPFS/%': '', '12mPFS/%': '', 'DFS/%': '', '证据ID': 'E009', '证据等级': 'A', '备注': 'ORR提升但OS未提升'},
    {'临床试验': 'NOTABLE', 'mOS/月': '10.9 vs 8.5', 'OS_HR': '0.66', '1yOS/%': '43.6 vs 26.8', '5yOS/%': '', 'mPFS/月': '4.2 vs 3.6', '6mPFS/%': '', '12mPFS/%': '', 'DFS/%': '', '证据ID': 'E010', '证据等级': 'A', '备注': ''},
    {'临床试验': 'HALO-301', 'mOS/月': '11.2 vs 11.5', 'OS_HR': '', '1yOS/%': '', '5yOS/%': '', 'mPFS/月': '7.1 vs 7.1', '6mPFS/%': '', '12mPFS/%': '', 'DFS/%': '', '证据ID': 'E011', '证据等级': 'A', '备注': 'ORR提升但OS/PFS无改善'},
    {'临床试验': 'PA.3', 'mOS/月': '6.24 vs 5.91', 'OS_HR': '0.82', '1yOS/%': '23 vs 17', '5yOS/%': '', 'mPFS/月': '3.75 vs 3.55', '6mPFS/%': '', '12mPFS/%': '', 'DFS/%': '', '证据ID': 'E012', '证据等级': 'A', '备注': ''},
    {'临床试验': 'TeloVac', 'mOS/月': '7.9 vs 6.9 vs 8.4', 'OS_HR': '', '1yOS/%': '33.7 vs 25.3 vs 32.3', '5yOS/%': '', 'mPFS/月': '', '6mPFS/%': '', '12mPFS/%': '', 'DFS/%': '', '证据ID': 'E013', '证据等级': 'C', '备注': '全文下载受限，主要为摘要级证据'},
    {'临床试验': 'KG4/2015', 'mOS/月': '11.3 vs 7.5', 'OS_HR': '', '1yOS/%': '', '5yOS/%': '', 'mPFS/月': '7.3 vs 4.6', '6mPFS/%': '', '12mPFS/%': '', 'DFS/%': '', '证据ID': 'E014', '证据等级': 'A', '备注': 'Eotaxin高表达亚组'},
    {'临床试验': 'CONKO-007', 'mOS/月': '14 vs 15', 'OS_HR': '0.94', '1yOS/%': '', '5yOS/%': '', 'mPFS/月': '', '6mPFS/%': '', '12mPFS/%': '', 'DFS/%': '8 vs 9', '证据ID': 'E015;E025', '证据等级': 'A/C', '备注': '随机主比较OS无差异；手术亚组获益'},
    {'临床试验': 'LAP-07', 'mOS/月': '16.5 vs 15.2', 'OS_HR': '', '1yOS/%': '', '5yOS/%': '', 'mPFS/月': '8.4 vs 9.9', '6mPFS/%': '', '12mPFS/%': '', 'DFS/%': '', '证据ID': 'E016', '证据等级': 'A', '备注': '局部控制改善，OS未改善'},
    {'临床试验': 'ACCORD11_QOL', 'mOS/月': '', 'OS_HR': '', '1yOS/%': '', '5yOS/%': '', 'mPFS/月': '', '6mPFS/%': '', '12mPFS/%': '', 'DFS/%': '', '证据ID': 'E017', '证据等级': 'B', '备注': 'QoL伴随论文，生存细节沿用主试验'},
    {'临床试验': 'PA3_QOL', 'mOS/月': '', 'OS_HR': '', '1yOS/%': '', '5yOS/%': '', 'mPFS/月': '', '6mPFS/%': '', '12mPFS/%': '', 'DFS/%': '', '证据ID': 'E018', '证据等级': 'B', '备注': 'QoL预后分析'},
    {'临床试验': 'MPACT_HRQoL', 'mOS/月': '', 'OS_HR': '', '1yOS/%': '', '5yOS/%': '', 'mPFS/月': '', '6mPFS/%': '', '12mPFS/%': '', 'DFS/%': '', '证据ID': 'E019', '证据等级': 'B', '备注': '真实世界QoL论文'},
    {'临床试验': 'MPACT_NICE_CE', 'mOS/月': '', 'OS_HR': '', '1yOS/%': '', '5yOS/%': '', 'mPFS/月': '', '6mPFS/%': '', '12mPFS/%': '', 'DFS/%': '', '证据ID': 'E020', '证据等级': 'B', '备注': '经济学证据，非随机对照主文'},
    {'临床试验': 'APICE_2018', 'mOS/月': '', 'OS_HR': '', '1yOS/%': '', '5yOS/%': '', 'mPFS/月': '', '6mPFS/%': '', '12mPFS/%': '', 'DFS/%': '', '证据ID': 'E021', '证据等级': 'B', '备注': '成本效果分析'},
    {'临床试验': 'FOLFIRINOX_vs_GNP_CE', 'mOS/月': '', 'OS_HR': '', '1yOS/%': '', '5yOS/%': '', 'mPFS/月': '', '6mPFS/%': '', '12mPFS/%': '', 'DFS/%': '', '证据ID': 'E022', '证据等级': 'B', '备注': '成本效果分析'},
    {'临床试验': 'LAPACT', 'mOS/月': '', 'OS_HR': '', '1yOS/%': '', '5yOS/%': '', 'mPFS/月': '', '6mPFS/%': '', '12mPFS/%': '', 'DFS/%': '', '证据ID': 'E023', '证据等级': 'B', '备注': 'II期研究，主终点TTF'},
]

TASK2_ROWS_RAW = [
    {'临床试验': 'PANOVA-3', 'ORR/%': '', 'CR/%': '', 'DCR/%': '', 'TTP/月': '', 'DOR/月': '', '手术率/%': '5.9 vs 4.7', 'R0切除率/%': '', 'pCR/%': '', '局部复发率/%': '', '远处转移率/%': '', 'CBR/%': '', '其他肿瘤控制指标': '摘要显示含ORR与切除相关终点', '证据ID': 'E001', '证据等级': 'B/C', '备注': ''},
    {'临床试验': 'ACCORD 11', 'ORR/%': '31.6 vs 9.4', 'CR/%': '', 'DCR/%': '70.2 vs 50.9', 'TTP/月': '', 'DOR/月': '5.9 vs 3.9', '手术率/%': '', 'R0切除率/%': '', 'pCR/%': '', '局部复发率/%': '', '远处转移率/%': '', 'CBR/%': '', '其他肿瘤控制指标': '', '证据ID': 'E002', '证据等级': 'A', '备注': ''},
    {'临床试验': 'MPACT', 'ORR/%': '23 vs 7', 'CR/%': '', 'DCR/%': '48 vs 33', 'TTP/月': '', 'DOR/月': '', '手术率/%': '', 'R0切除率/%': '', 'pCR/%': '', '局部复发率/%': '', '远处转移率/%': '', 'CBR/%': '', '其他肿瘤控制指标': 'CA19-9应答率更高', '证据ID': 'E003', '证据等级': 'A', '备注': ''},
    {'临床试验': 'NEOPAN', 'ORR/%': '42.4 vs 15.1', 'CR/%': '8 vs 2.3', 'DCR/%': '', 'TTP/月': '', 'DOR/月': '', '手术率/%': '5.9 vs 4.7', 'R0切除率/%': '5.9 vs 4.7', 'pCR/%': '', '局部复发率/%': '', '远处转移率/%': '', 'CBR/%': '', '其他肿瘤控制指标': '', '证据ID': 'E004', '证据等级': 'A', '备注': ''},
    {'临床试验': 'NAPOLI-3', 'ORR/%': '41.8 vs 36.2', 'CR/%': '<1 vs <1', 'DCR/%': '69 vs 73', 'TTP/月': '', 'DOR/月': '7.3 vs 5.0', '手术率/%': '', 'R0切除率/%': '', 'pCR/%': '', '局部复发率/%': '', '远处转移率/%': '', 'CBR/%': '', '其他肿瘤控制指标': '', '证据ID': 'E005', '证据等级': 'A', '备注': ''},
    {'临床试验': 'JCOG1611 (GENERATE)', 'ORR/%': '32.4 vs 42.4 vs 35.4', 'CR/%': '0.6 vs 0 vs 0', 'DCR/%': '72.9 vs 81.8 vs 83.4', 'TTP/月': '', 'DOR/月': '', '手术率/%': '', 'R0切除率/%': '', 'pCR/%': '', '局部复发率/%': '', '远处转移率/%': '', 'CBR/%': '', '其他肿瘤控制指标': '三臂二线转化潜力差异', '证据ID': 'E006', '证据等级': 'A', '备注': ''},
    {'临床试验': 'GEST', 'ORR/%': '13.5 vs 8.0 vs 19.0', 'CR/%': '0.4 vs 0.5 vs 0.8', 'DCR/%': '62.7 vs 63.3 vs 75.1', 'TTP/月': '', 'DOR/月': '', '手术率/%': '', 'R0切除率/%': '', 'pCR/%': '', '局部复发率/%': '', '远处转移率/%': '', 'CBR/%': '', '其他肿瘤控制指标': 'GS组QAPFM更长', '证据ID': 'E007;E024', '证据等级': 'A/C', '备注': ''},
    {'临床试验': 'GEST++(IPD)', 'ORR/%': '8.3 vs 29.6', 'CR/%': '0 vs 1.2', 'DCR/%': '77.4 vs 82.7', 'TTP/月': '', 'DOR/月': '', '手术率/%': '', 'R0切除率/%': '', 'pCR/%': '', '局部复发率/%': '', '远处转移率/%': '', 'CBR/%': '', '其他肿瘤控制指标': 'LAPC亚组肿瘤缩小更明显', '证据ID': 'E008', '证据等级': 'A', '备注': ''},
    {'临床试验': 'AVENGER 500', 'ORR/%': '39.1 vs 34.4', 'CR/%': '', 'DCR/%': '', 'TTP/月': '', 'DOR/月': '', '手术率/%': '', 'R0切除率/%': '', 'pCR/%': '', '局部复发率/%': '', '远处转移率/%': '', 'CBR/%': '', '其他肿瘤控制指标': 'ORR改善但OS不改善', '证据ID': 'E009', '证据等级': 'A', '备注': ''},
    {'临床试验': 'NOTABLE', 'ORR/%': '7.3 vs 9.8', 'CR/%': '0 vs 0', 'DCR/%': '68.3 vs 63.4', 'TTP/月': '', 'DOR/月': '', '手术率/%': '', 'R0切除率/%': '', 'pCR/%': '', '局部复发率/%': '', '远处转移率/%': '', 'CBR/%': '39.3 vs 32.2', '其他肿瘤控制指标': '', '证据ID': 'E010', '证据等级': 'A', '备注': 'CBR按Burris标准'},
    {'临床试验': 'HALO-301', 'ORR/%': '47 vs 36', 'CR/%': '', 'DCR/%': '', 'TTP/月': '', 'DOR/月': '6.1 vs 7.4', '手术率/%': '', 'R0切除率/%': '', 'pCR/%': '', '局部复发率/%': '', '远处转移率/%': '', 'CBR/%': '', '其他肿瘤控制指标': '高HA亚组探索性获益', '证据ID': 'E011', '证据等级': 'A', '备注': ''},
    {'临床试验': 'PA.3', 'ORR/%': '8.6 vs 8.0', 'CR/%': '', 'DCR/%': '57.5 vs 49.2', 'TTP/月': '', 'DOR/月': '163天 vs 163天', '手术率/%': '', 'R0切除率/%': '', 'pCR/%': '', '局部复发率/%': '', '远处转移率/%': '', 'CBR/%': '', '其他肿瘤控制指标': '', '证据ID': 'E012', '证据等级': 'A', '备注': ''},
    {'临床试验': 'TeloVac', 'ORR/%': '', 'CR/%': '', 'DCR/%': '', 'TTP/月': '', 'DOR/月': '', '手术率/%': '', 'R0切除率/%': '', 'pCR/%': '', '局部复发率/%': '', '远处转移率/%': '', 'CBR/%': '', '其他肿瘤控制指标': '公开摘要未报告可核数值', '证据ID': 'E013', '证据等级': 'C', '备注': ''},
    {'临床试验': 'KG4/2015', 'ORR/%': '26.7 vs 27.4', 'CR/%': '0 vs 0', 'DCR/%': '72.0 vs 63.0', 'TTP/月': '7.3 vs 4.5', 'DOR/月': '', '手术率/%': '', 'R0切除率/%': '', 'pCR/%': '', '局部复发率/%': '', '远处转移率/%': '', 'CBR/%': '', '其他肿瘤控制指标': '免疫诱导后生物标志物相关获益', '证据ID': 'E014', '证据等级': 'A', '备注': ''},
    {'临床试验': 'CONKO-007', 'ORR/%': '', 'CR/%': '', 'DCR/%': '', 'TTP/月': '', 'DOR/月': '', '手术率/%': '76 vs 75', 'R0切除率/%': '25 vs 18 (ITT); 69.4 vs 50.0 (手术人群)', 'pCR/%': '1 vs 2', '局部复发率/%': '46.9 vs 63.2', '远处转移率/%': '82.7 vs 84.7', 'CBR/%': '', '其他肿瘤控制指标': '手术vs未手术 mOS 19 vs 13', '证据ID': 'E015;E025', '证据等级': 'A/C', '备注': ''},
    {'临床试验': 'LAP-07', 'ORR/%': '', 'CR/%': '', 'DCR/%': '', 'TTP/月': '', 'DOR/月': '', '手术率/%': '', 'R0切除率/%': '', 'pCR/%': '', '局部复发率/%': '32 vs 46', '远处转移率/%': '', 'CBR/%': '', '其他肿瘤控制指标': '局部进展风险下降', '证据ID': 'E016', '证据等级': 'A', '备注': ''},
    {'临床试验': 'ACCORD11_QOL', 'ORR/%': '', 'CR/%': '', 'DCR/%': '', 'TTP/月': '', 'DOR/月': '', '手术率/%': '', 'R0切除率/%': '', 'pCR/%': '', '局部复发率/%': '', '远处转移率/%': '', 'CBR/%': '', '其他肿瘤控制指标': 'QoL恶化时间显著延后', '证据ID': 'E017', '证据等级': 'B', '备注': ''},
    {'临床试验': 'PA3_QOL', 'ORR/%': '', 'CR/%': '', 'DCR/%': '', 'TTP/月': '', 'DOR/月': '', '手术率/%': '', 'R0切除率/%': '', 'pCR/%': '', '局部复发率/%': '', '远处转移率/%': '', 'CBR/%': '', '其他肿瘤控制指标': 'QoL基线与8周变化可预测OS', '证据ID': 'E018', '证据等级': 'B', '备注': ''},
    {'临床试验': 'MPACT_HRQoL', 'ORR/%': '', 'CR/%': '', 'DCR/%': '', 'TTP/月': '', 'DOR/月': '', '手术率/%': '', 'R0切除率/%': '', 'pCR/%': '', '局部复发率/%': '', '远处转移率/%': '', 'CBR/%': '', '其他肿瘤控制指标': '真实世界QoL横断评估', '证据ID': 'E019', '证据等级': 'B', '备注': ''},
    {'临床试验': 'MPACT_NICE_CE', 'ORR/%': '', 'CR/%': '', 'DCR/%': '', 'TTP/月': '', 'DOR/月': '', '手术率/%': '', 'R0切除率/%': '', 'pCR/%': '', '局部复发率/%': '', '远处转移率/%': '', 'CBR/%': '', '其他肿瘤控制指标': 'CA046显示OS/PFS优于Gem', '证据ID': 'E020', '证据等级': 'B', '备注': ''},
    {'临床试验': 'APICE_2018', 'ORR/%': '', 'CR/%': '', 'DCR/%': '', 'TTP/月': '', 'DOR/月': '', '手术率/%': '', 'R0切除率/%': '', 'pCR/%': '', '局部复发率/%': '', '远处转移率/%': '', 'CBR/%': '', '其他肿瘤控制指标': 'Markov模型: Nab-P+G增益0.154 QALY', '证据ID': 'E021', '证据等级': 'B', '备注': ''},
    {'临床试验': 'FOLFIRINOX_vs_GNP_CE', 'ORR/%': '', 'CR/%': '', 'DCR/%': '', 'TTP/月': '', 'DOR/月': '', '手术率/%': '', 'R0切除率/%': '', 'pCR/%': '', '局部复发率/%': '', '远处转移率/%': '', 'CBR/%': '', '其他肿瘤控制指标': '模型QALY: FOLFIRINOX 0.67 vs GEM-N 0.51', '证据ID': 'E022', '证据等级': 'B', '备注': ''},
    {'临床试验': 'LAPACT', 'ORR/%': '', 'CR/%': '', 'DCR/%': '', 'TTP/月': '主要终点', 'DOR/月': '', '手术率/%': '', 'R0切除率/%': '', 'pCR/%': '', '局部复发率/%': '', '远处转移率/%': '', 'CBR/%': '', '其他肿瘤控制指标': '次要终点含ORR/DCR/PFS/OS', '证据ID': 'E023', '证据等级': 'B', '备注': ''},
]

TASK3_ROWS_RAW = [
    {'临床试验': 'PANOVA-3', '≥3级AE/%': '', '治疗相关死亡/%': '', 'QOL': '方案设计包含QOL作为关键次要终点', 'QALY/QALM': '', '痛苦评分/症状': '包含pain-free survival', 'PRO工具': '', '关键AE信号': '', '证据ID': 'E001', '证据等级': 'B/C', '备注': '会议阶段未公开完整AE矩阵'},
    {'临床试验': 'ACCORD 11', '≥3级AE/%': '明确报告grade3/4毒性', '治疗相关死亡/%': '2例', 'QOL': '6月时QoL明确恶化 31% vs 66%', 'QALY/QALM': '', '痛苦评分/症状': '痛苦维度TUDD显著延后', 'PRO工具': 'EORTC QLQ-C30', '关键AE信号': '发热性中性粒细胞减少 5.4%', '证据ID': 'E002', '证据等级': 'A', '备注': ''},
    {'临床试验': 'MPACT', '≥3级AE/%': '中性粒细胞减少 38% vs 27%', '治疗相关死亡/%': '', 'QOL': '主文未报告完整患者量表结局', 'QALY/QALM': '', '痛苦评分/症状': '', 'PRO工具': '', '关键AE信号': '疲劳 17% vs 7%; 神经病变 17% vs 1%', '证据ID': 'E003', '证据等级': 'A', '备注': ''},
    {'临床试验': 'NEOPAN', '≥3级AE/%': 'grade3/4 serious AE 41% vs 32%', '治疗相关死亡/%': '', 'QOL': '无显著QOL恶化', 'QALY/QALM': '', '痛苦评分/症状': '', 'PRO工具': 'EORTC QLQ-C30', '关键AE信号': 'PFS获益伴中度毒性增加', '证据ID': 'E004', '证据等级': 'A', '备注': ''},
    {'临床试验': 'NAPOLI-3', '≥3级AE/%': '87% vs 86%', '治疗相关死亡/%': '2% vs 2%', 'QOL': '主文未给出QALY', 'QALY/QALM': '', '痛苦评分/症状': '', 'PRO工具': '', '关键AE信号': 'grade3-4中性粒细胞减少/腹泻更常见', '证据ID': 'E005', '证据等级': 'A', '备注': ''},
    {'临床试验': 'JCOG1611 (GENERATE)', '≥3级AE/%': '厌食 23.3%/27.5% vs 5.0%', '治疗相关死亡/%': '0.2% (S-IROX组1例)', 'QOL': '未设主要QOL终点', 'QALY/QALM': '', '痛苦评分/症状': '', 'PRO工具': '', '关键AE信号': '两强化方案毒性更高', '证据ID': 'E006', '证据等级': 'A', '备注': ''},
    {'临床试验': 'GEST', '≥3级AE/%': '按CTCAE评估，GS毒性略高', '治疗相关死亡/%': '', 'QOL': 'GS HRQOL优于Gem', 'QALY/QALM': 'QALMs/QAPFMs显著更长', '痛苦评分/症状': 'fatigue/anorexia/pain与EQ-5D相关', 'PRO工具': 'EQ-5D', '关键AE信号': '毒性-获益平衡可接受', '证据ID': 'E007', '证据等级': 'A', '备注': '任务3核心高价值证据'},
    {'临床试验': 'GEST++(IPD)', '≥3级AE/%': 'GS组grade≥3毒性更高(特定项)', '治疗相关死亡/%': '', 'QOL': '', 'QALY/QALM': '', '痛苦评分/症状': '', 'PRO工具': '', '关键AE信号': 'rash/腹泻/中性粒细胞减少差异', '证据ID': 'E008', '证据等级': 'A', '备注': ''},
    {'临床试验': 'AVENGER 500', '≥3级AE/%': 'grade3 57.9% vs 48.9%; grade4 22.0% vs 28.5%', '治疗相关死亡/%': '', 'QOL': '', 'QALY/QALM': '', '痛苦评分/症状': '', 'PRO工具': '', '关键AE信号': '毒性更高但OS未获益', '证据ID': 'E009', '证据等级': 'A', '备注': ''},
    {'临床试验': 'NOTABLE', '≥3级AE/%': '无grade4-5 AE', '治疗相关死亡/%': '0', 'QOL': '', 'QALY/QALM': '', '痛苦评分/症状': '', 'PRO工具': '', '关键AE信号': 'TRAE主要1-3级', '证据ID': 'E010', '证据等级': 'A', '备注': ''},
    {'临床试验': 'HALO-301', '≥3级AE/%': '疲劳 16.0% vs 9.6%', '治疗相关死亡/%': '', 'QOL': '', 'QALY/QALM': '', '痛苦评分/症状': '', 'PRO工具': '', '关键AE信号': 'AE增加但OS/PFS不变', '证据ID': 'E011', '证据等级': 'A', '备注': ''},
    {'临床试验': 'PA.3', '≥3级AE/%': '中性粒细胞减少 24% vs 27%', '治疗相关死亡/%': '方案相关死亡存在', 'QOL': 'QOL已系统评估', 'QALY/QALM': '', '痛苦评分/症状': '基线pain score参与分层', 'PRO工具': 'EORTC QLQ-C30', '关键AE信号': '皮疹/腹泻增加，血液学毒性相近', '证据ID': 'E012', '证据等级': 'A', '备注': ''},
    {'临床试验': 'TeloVac', '≥3级AE/%': '', '治疗相关死亡/%': '', 'QOL': '仅摘要级可得', 'QALY/QALM': '', '痛苦评分/症状': '', 'PRO工具': '', '关键AE信号': '', '证据ID': 'E013', '证据等级': 'C', '备注': '待补全文'},
    {'临床试验': 'KG4/2015', '≥3级AE/%': '77.3% vs 73.1%', '治疗相关死亡/%': '', 'QOL': '两组QOL量表趋势相近', 'QALY/QALM': '报告EQ-5D-5L', '痛苦评分/症状': 'pain/discomfort条目纳入', 'PRO工具': 'EORTC QLQ-C30 + EQ-5D-5L', '关键AE信号': '血液学AE为主', '证据ID': 'E014', '证据等级': 'A', '备注': ''},
    {'临床试验': 'CONKO-007', '≥3级AE/%': 'CRT组白细胞减少/血小板减少更高', '治疗相关死亡/%': '有grade5毒性报告', 'QOL': '', 'QALY/QALM': '', '痛苦评分/症状': '', 'PRO工具': '', '关键AE信号': '局控改善但毒性结构改变', '证据ID': 'E015', '证据等级': 'A', '备注': ''},
    {'临床试验': 'LAP-07', '≥3级AE/%': '总体未增加(除恶心)', '治疗相关死亡/%': '', 'QOL': '', 'QALY/QALM': '', '痛苦评分/症状': '', 'PRO工具': '', '关键AE信号': '局部进展下降伴毒性总体可控', '证据ID': 'E016', '证据等级': 'A', '备注': ''},
    {'临床试验': 'ACCORD11_QOL', '≥3级AE/%': '伴随论文聚焦QoL，安全性沿主试验', '治疗相关死亡/%': '', 'QOL': 'GHS/功能域TUDD显著延后', 'QALY/QALM': '', '痛苦评分/症状': '疼痛/失眠/食欲等域改善', 'PRO工具': 'EORTC QLQ-C30', '关键AE信号': '腹泻在前2月显著增加', '证据ID': 'E017', '证据等级': 'B', '备注': ''},
    {'临床试验': 'PA3_QOL', '≥3级AE/%': '', '治疗相关死亡/%': '', 'QOL': '基线PF越高OS越长 (HR 0.86)', 'QALY/QALM': '', '痛苦评分/症状': '8周PF提升预测OS改善 (HR 0.89)', 'PRO工具': 'EORTC QLQ-C30', '关键AE信号': '', '证据ID': 'E018', '证据等级': 'B', '备注': ''},
    {'临床试验': 'MPACT_HRQoL', '≥3级AE/%': '', '治疗相关死亡/%': '', 'QOL': '真实世界mPC患者QoL在治疗阶段间差异明显', 'QALY/QALM': '', '痛苦评分/症状': '晚期病程症状负担高', 'PRO工具': 'EORTC QLQ-C30/PAN26/EQ-5D', '关键AE信号': '', '证据ID': 'E019', '证据等级': 'B', '备注': ''},
    {'临床试验': 'MPACT_NICE_CE', '≥3级AE/%': '模型考虑AE disutility', '治疗相关死亡/%': '', 'QOL': 'NICE复评纳入HRQoL证据', 'QALY/QALM': 'ICER £41,000-£46,000 /QALY (vs Gem)', '痛苦评分/症状': '', 'PRO工具': 'EQ-5D映射', '关键AE信号': 'AE双计数风险被ERG指出', '证据ID': 'E020', '证据等级': 'B', '备注': '来源PMC BioC全文XML'},
    {'临床试验': 'APICE_2018', '≥3级AE/%': '', '治疗相关死亡/%': '', 'QOL': '', 'QALY/QALM': '增量QALY 0.154; ICUR €46,021.58/QALY', '痛苦评分/症状': '', 'PRO工具': '成本效用模型', '关键AE信号': '', '证据ID': 'E021', '证据等级': 'B', '备注': ''},
    {'临床试验': 'FOLFIRINOX_vs_GNP_CE', '≥3级AE/%': '', '治疗相关死亡/%': '', 'QOL': '', 'QALY/QALM': 'QALY: 0.67 vs 0.51; ICER $32,019.75/QALY', '痛苦评分/症状': '', 'PRO工具': '决策模型', '关键AE信号': '', '证据ID': 'E022', '证据等级': 'B', '备注': ''},
    {'临床试验': 'LAPACT', '≥3级AE/%': '安全性为次要终点', '治疗相关死亡/%': '', 'QOL': 'QoL为次要终点', 'QALY/QALM': '', '痛苦评分/症状': '', 'PRO工具': '', '关键AE信号': '', '证据ID': 'E023', '证据等级': 'B', '备注': '前瞻性II期'},
]

EVIDENCE_SPECS = [
    {
        'datum_id': 'E001', 'task_id': '任务1/2/3', 'trial_id': 'PANOVA-3', 'endpoint': 'OS/PFS与QOL框架',
        'value_text': 'mOS 16.2 vs 14.2; mPFS 10.6 vs 9.3; 1yOS 68.1 vs 60.2; 12mPFS 43.9 vs 34.1',
        'population': 'LAPC', 'arm_compare': 'TTFields+GnP vs GnP', 'source_type': 'user_screenshot',
        'doi/pmid': '', 'file_path': 'user_attachment_images', 'search_phrase': None,
        'manual_quote': '用户截图显示PANOVA-3行包含OS/PFS/1yOS/12mPFS与手术率字段，且试验设计包含QOL与安全性终点。', 'manual_page': '截图', 'evidence_level': 'B/C',
    },
    {
        'datum_id': 'E002', 'task_id': '任务1/2/3', 'trial_id': 'ACCORD 11', 'endpoint': 'OS/PFS/ORR/AE/QOL',
        'value_text': 'mOS 11.1 vs 6.8; mPFS 6.4 vs 3.3; ORR 31.6 vs 9.4; QoL恶化31% vs 66%',
        'population': 'MPC', 'arm_compare': 'FOLFIRINOX vs Gem', 'source_type': 'fulltext_pdf',
        'doi/pmid': '10.1056/NEJMoa1011923 / 21561347', 'file_path': str(PDF_MAP['ACCORD 11']),
        'search_phrase': 'At 6 months, 31% of the patients in the FOLFIRINOX group had a definitive degradation of the quality of life versus 66%',
        'manual_quote': None, 'manual_page': None, 'evidence_level': 'A',
    },
    {
        'datum_id': 'E003', 'task_id': '任务1/2/3', 'trial_id': 'MPACT', 'endpoint': 'OS/PFS/ORR/关键AE',
        'value_text': 'mOS 8.5 vs 6.7; mPFS 5.5 vs 3.7; ORR 23 vs 7; neutropenia 38% vs 27%',
        'population': 'MPC', 'arm_compare': 'GnP vs Gem', 'source_type': 'fulltext_pdf',
        'doi/pmid': '10.1056/NEJMoa1304369 / 24131140', 'file_path': str(PDF_MAP['MPACT']),
        'search_phrase': 'The most common adverse events of grade 3 or higher were neutropenia (38% in the nab-paclitaxel–gemcitabine group vs. 27% in the gemcitabine group)',
        'manual_quote': None, 'manual_page': None, 'evidence_level': 'A',
    },
    {
        'datum_id': 'E004', 'task_id': '任务1/2/3', 'trial_id': 'NEOPAN', 'endpoint': 'PFS/OS/ORR/AE/QOL',
        'value_text': 'mPFS 9.7 vs 7.7; mOS 15.7 vs 15.4; ORR 42.4 vs 15.1; serious grade3/4 AE 41% vs 32%',
        'population': 'LAPC', 'arm_compare': 'mFOLFIRINOX vs Gem', 'source_type': 'fulltext_pdf',
        'doi/pmid': '10.1200/JCO-24-02210 / 40378359', 'file_path': str(PDF_MAP['NEOPAN']),
        'search_phrase': 'no deterioration in quality of life',
        'manual_quote': None, 'manual_page': None, 'evidence_level': 'A',
    },
    {
        'datum_id': 'E005', 'task_id': '任务1/2/3', 'trial_id': 'NAPOLI-3', 'endpoint': 'OS/PFS/ORR/AE/死亡',
        'value_text': 'mOS 11.1 vs 9.2; mPFS 7.4 vs 5.6; grade≥3 AE 87% vs 86%; treatment-related death 2% vs 2%',
        'population': 'MPC', 'arm_compare': 'NALIRIFOX vs GnP', 'source_type': 'fulltext_pdf',
        'doi/pmid': '10.1016/S0140-6736(23)00930-2 / 36990693', 'file_path': str(PDF_MAP['NAPOLI-3']),
        'search_phrase': 'Grade 3 or higher treatment-emergent adverse events occurred in 322 (87%) of 370 patients receiving NALIRIFOX and 326 (86%)',
        'manual_quote': None, 'manual_page': None, 'evidence_level': 'A',
    },
    {
        'datum_id': 'E006', 'task_id': '任务1/2/3', 'trial_id': 'JCOG1611 (GENERATE)', 'endpoint': 'OS/PFS/ORR/关键AE',
        'value_text': 'mOS 14.0/13.6/17.0; mPFS 5.8/6.7/6.7; anorexia grade3/4 23.3%/27.5% vs 5.0%',
        'population': 'MPC/rPC', 'arm_compare': 'mFOLFIRINOX vs S-IROX vs GnP', 'source_type': 'fulltext_pdf',
        'doi/pmid': '', 'file_path': str(PDF_MAP['JCOG1611 (GENERATE)']),
        'search_phrase': 'Grade 3 to 4 anorexia was more frequent in the mFOLFIRINOX (23.3%) and S-IROX (27.5%) groups than in the nab-paclitaxel 1 gemcitabine group (5.0%)',
        'manual_quote': None, 'manual_page': None, 'evidence_level': 'A',
    },
    {
        'datum_id': 'E007', 'task_id': '任务1/2/3', 'trial_id': 'GEST', 'endpoint': 'HRQOL/QALM/QAPFM',
        'value_text': 'GS组QALMs/QAPFMs/TUDD显著优于Gem',
        'population': 'LAPC/MPC', 'arm_compare': 'Gem vs S-1 vs GS', 'source_type': 'fulltext_pdf',
        'doi/pmid': '10.1136/esmoopen-2016-000151 / 28761731', 'file_path': str(PDF_MAP['GEST']),
        'search_phrase': 'QALMs, QAPFMs and TUDD were significantly longer in the GS than gemcitabine group',
        'manual_quote': None, 'manual_page': None, 'evidence_level': 'A',
    },
    {
        'datum_id': 'E008', 'task_id': '任务1/2/3', 'trial_id': 'GEST++(IPD)', 'endpoint': 'LAPC亚组疗效与毒性',
        'value_text': 'mOS 11.83 vs 16.41; mPFS 5.78 vs 11.76; grade≥3毒性在GS组更高',
        'population': 'LAPC', 'arm_compare': 'Gem vs GS', 'source_type': 'fulltext_pdf',
        'doi/pmid': '10.1038/bjc.2017.118 / 28427067', 'file_path': str(PDF_MAP['GEST++(IPD)']),
        'search_phrase': 'The incidences of the following grade 3 or higher AEs were significantly higher in the GS group',
        'manual_quote': None, 'manual_page': None, 'evidence_level': 'A',
    },
    {
        'datum_id': 'E009', 'task_id': '任务1/2/3', 'trial_id': 'AVENGER 500', 'endpoint': 'OS/PFS/ORR与分级AE',
        'value_text': 'mOS 11.10 vs 11.73; mPFS 7.82 vs 7.98; ORR 39.1 vs 34.4; grade3/4 AE差异',
        'population': 'MPC', 'arm_compare': 'mFFX+Devimistat vs FFX', 'source_type': 'fulltext_pdf',
        'doi/pmid': '', 'file_path': str(PDF_MAP['AVENGER 500']),
        'search_phrase': 'frequency of CTCAE grade 3 events',
        'manual_quote': None, 'manual_page': None, 'evidence_level': 'A',
    },
    {
        'datum_id': 'E010', 'task_id': '任务1/2/3', 'trial_id': 'NOTABLE', 'endpoint': 'OS/PFS/CBR与安全性',
        'value_text': 'mOS 10.9 vs 8.5; mPFS 4.2 vs 3.6; CBR 39.3 vs 32.2; 无grade4-5 AE',
        'population': 'KRAS野生型 LAPC/MPC', 'arm_compare': 'Gem+尼妥珠单抗 vs Gem', 'source_type': 'fulltext_pdf',
        'doi/pmid': '', 'file_path': str(PDF_MAP['NOTABLE']),
        'search_phrase': 'No grade 4-5 AEs were observed',
        'manual_quote': None, 'manual_page': None, 'evidence_level': 'A',
    },
    {
        'datum_id': 'E011', 'task_id': '任务1/2/3', 'trial_id': 'HALO-301', 'endpoint': 'ORR提升但OS/PFS无改善',
        'value_text': 'ORR 47 vs 36; mOS 11.2 vs 11.5; mPFS 7.1 vs 7.1; fatigue grade≥3 16.0% vs 9.6%',
        'population': 'MPC', 'arm_compare': 'GnP+PEGPH20 vs GnP', 'source_type': 'fulltext_pdf',
        'doi/pmid': '10.1200/JCO.20.00590 / 32673165', 'file_path': str(PDF_MAP['HALO-301']),
        'search_phrase': 'Grade $ 3 adverse events',
        'manual_quote': None, 'manual_page': None, 'evidence_level': 'A',
    },
    {
        'datum_id': 'E012', 'task_id': '任务1/2/3', 'trial_id': 'PA.3', 'endpoint': 'OS/PFS与毒性/QOL',
        'value_text': 'mOS 6.24 vs 5.91; mPFS 3.75 vs 3.55; grade3/4 neutropenia 24% vs 27%',
        'population': 'LAPC/MPC', 'arm_compare': 'Gem+Erlotinib vs Gem', 'source_type': 'fulltext_pdf',
        'doi/pmid': '10.1200/JCO.2006.07.9525 / 17452677', 'file_path': str(PDF_MAP['PA.3']),
        'search_phrase': 'grade 3/4 neutropenia and thrombocytopenia seen in 24%',
        'manual_quote': None, 'manual_page': None, 'evidence_level': 'A',
    },
    {
        'datum_id': 'E013', 'task_id': '任务1/2/3', 'trial_id': 'TeloVac', 'endpoint': '免疫治疗试验摘要证据',
        'value_text': '研究聚焦疗效与安全性，数值未完整公开',
        'population': 'LAPC/MPC', 'arm_compare': 'GemCap方案±GV1001', 'source_type': 'pubmed_abstract',
        'doi/pmid': '10.1016/S1470-2045(14)70236-0 / 24954781', 'file_path': 'https://pubmed.ncbi.nlm.nih.gov/24954781/',
        'search_phrase': None,
        'manual_quote': 'We aimed to assess the efficacy and safety of sequential or simultaneous telomerase vaccination (GV1001) in combination with chemotherapy.',
        'manual_page': 'Abstract', 'evidence_level': 'C',
    },
    {
        'datum_id': 'E014', 'task_id': '任务1/2/3', 'trial_id': 'KG4/2015', 'endpoint': 'OS/TTP/AE/QOL',
        'value_text': 'mOS 11.3 vs 7.5; TTP 7.3 vs 4.5; grade≥3 AE 77.3% vs 73.1%',
        'population': 'Eotaxin高表达 LAPC/MPC', 'arm_compare': 'GemCap+GV1001 vs GemCap', 'source_type': 'fulltext_pdf',
        'doi/pmid': '10.1038/s41416-023-02474-w / 38278089', 'file_path': str(PDF_MAP['KG4/2015']),
        'search_phrase': 'AEs ≥grade 3 were reported in 58 cases (77.3%) and 49 cases (73.1%)',
        'manual_quote': None, 'manual_page': None, 'evidence_level': 'A',
    },
    {
        'datum_id': 'E015', 'task_id': '任务1/2/3', 'trial_id': 'CONKO-007', 'endpoint': 'R0/局控/远转与grade3/4 AE',
        'value_text': 'R0 25% vs 18%(ITT); 局部复发46.9% vs 63.2%; 远处转移82.7% vs 84.7%',
        'population': 'LAPC', 'arm_compare': 'CRT vs ChT', 'source_type': 'fulltext_pdf',
        'doi/pmid': '10.1200/JCO-24-01502', 'file_path': str(PDF_MAP['CONKO-007']),
        'search_phrase': 'TABLE 3. Most Common Grade 3 or 4 Adverse Events by Treatment Regimen',
        'manual_quote': None, 'manual_page': None, 'evidence_level': 'A',
    },
    {
        'datum_id': 'E016', 'task_id': '任务1/2/3', 'trial_id': 'LAP-07', 'endpoint': '局部进展与毒性',
        'value_text': '局部进展32% vs 46%; 除恶心外未增加grade3-4毒性',
        'population': 'LAPC', 'arm_compare': 'CRT vs ChT', 'source_type': 'fulltext_pdf',
        'doi/pmid': '10.1001/jama.2016.4324 / 27139057', 'file_path': str(PDF_MAP['LAP-07']),
        'search_phrase': 'decreased local progression (32% vs 46%, P = .03) and no increase in grade 3 to 4 toxicity, except for nausea',
        'manual_quote': None, 'manual_page': None, 'evidence_level': 'A',
    },
    {
        'datum_id': 'E017', 'task_id': '任务3', 'trial_id': 'ACCORD11_QOL', 'endpoint': 'QoL伴随论文',
        'value_text': 'FOLFIRINOX在GHS/功能域TUDD显著延后，腹泻前2月增加',
        'population': 'MPC', 'arm_compare': 'FOLFIRINOX vs Gem', 'source_type': 'pubmed_abstract',
        'doi/pmid': '10.1200/JCO.2012.44.4869 / 23213101', 'file_path': 'https://pubmed.ncbi.nlm.nih.gov/23213101/',
        'search_phrase': None,
        'manual_quote': 'Time until definitive deterioration ≥20 points was significantly longer for FOLFIRINOX compared with gemcitabine for GHS and multiple functional/symptom domains.',
        'manual_page': 'Abstract', 'evidence_level': 'B',
    },
    {
        'datum_id': 'E018', 'task_id': '任务3', 'trial_id': 'PA3_QOL', 'endpoint': 'QoL预后价值',
        'value_text': '基线PF与8周PF改善均独立预测OS',
        'population': 'advanced pancreatic cancer', 'arm_compare': 'G+E vs G+P', 'source_type': 'pubmed_abstract',
        'doi/pmid': '10.1016/j.pan.2016.08.013 / 27600995', 'file_path': 'https://pubmed.ncbi.nlm.nih.gov/27600995/',
        'search_phrase': None,
        'manual_quote': 'Better baseline physical functioning predicted longer OS (HR 0.86), and week-8 PF improvement also predicted improved survival (HR 0.89 per 10-point increase).',
        'manual_page': 'Abstract', 'evidence_level': 'B',
    },
    {
        'datum_id': 'E019', 'task_id': '任务3', 'trial_id': 'MPACT_HRQoL', 'endpoint': '真实世界QoL',
        'value_text': '转移性胰腺癌不同治疗阶段QoL差异明显',
        'population': 'mPC real-world', 'arm_compare': 'observational', 'source_type': 'pubmed_abstract',
        'doi/pmid': '10.1007/s12029-016-9902-9 / 28028766', 'file_path': 'https://pubmed.ncbi.nlm.nih.gov/28028766/',
        'search_phrase': None,
        'manual_quote': 'This pilot project assessed real-world QoL in patients with mPC at different stages of treatment.',
        'manual_page': 'Abstract', 'evidence_level': 'B',
    },
    {
        'datum_id': 'E020', 'task_id': '任务2/3', 'trial_id': 'MPACT_NICE_CE', 'endpoint': 'QALY与ICER',
        'value_text': 'base-case ICER £46,932/QALY; plausible £41,000–£46,000/QALY',
        'population': 'untreated metastatic pancreatic cancer', 'arm_compare': 'Nab-Pac+Gem vs comparators', 'source_type': 'pmc_bioc_xml',
        'doi/pmid': '10.1007/s40273-018-0646-1 / 29600384', 'file_path': str(ROOT / 'downloads/new_papers/MPACT_NICE_CE_2018.xml'),
        'search_phrase': 'base-case ICER was £46,932 per quality-adjusted life-year (QALY) gained',
        'manual_quote': None, 'manual_page': 'XML-ABSTRACT', 'evidence_level': 'B',
    },
    {
        'datum_id': 'E021', 'task_id': '任务3', 'trial_id': 'APICE_2018', 'endpoint': '增量QALY与ICUR',
        'value_text': '增量QALY 0.154; ICUR €46,021.58/QALY',
        'population': 'MPC cost-effectiveness', 'arm_compare': 'Nab-P+G vs G', 'source_type': 'pubmed_abstract',
        'doi/pmid': '10.1080/14737167.2018.1464394 / 29641931', 'file_path': 'https://pubmed.ncbi.nlm.nih.gov/29641931/',
        'search_phrase': None,
        'manual_quote': 'Nab-P + G totals 0.154 incremental QALYs and €7082.68 incremental costs vs G alone; ICUR €46,021.58/QALY.',
        'manual_page': 'Abstract', 'evidence_level': 'B',
    },
    {
        'datum_id': 'E022', 'task_id': '任务3', 'trial_id': 'FOLFIRINOX_vs_GNP_CE', 'endpoint': 'QALY与ICER',
        'value_text': 'QALY 0.67 vs 0.51; ICER $32,019.75/QALY',
        'population': 'MPC cost-effectiveness', 'arm_compare': 'FOLFIRINOX vs GEM-N', 'source_type': 'pubmed_abstract',
        'doi/pmid': '10.5301/tj.5000499 / 27056335', 'file_path': 'https://pubmed.ncbi.nlm.nih.gov/27056335/',
        'search_phrase': None,
        'manual_quote': 'FOLFIRINOX yielded 0.67 QALY and GEM-N 0.51 QALY; ICER of FOLFIRINOX versus GEM-N was $32,019.75 per QALY gained.',
        'manual_page': 'Abstract', 'evidence_level': 'B',
    },
    {
        'datum_id': 'E023', 'task_id': '任务2/3', 'trial_id': 'LAPACT', 'endpoint': 'LAPC II期终点框架',
        'value_text': 'primary endpoint为TTF; secondary含DCR/ORR/PFS/OS/safety/QoL',
        'population': 'unresectable LAPC', 'arm_compare': 'single-arm induction', 'source_type': 'pubmed_abstract',
        'doi/pmid': '10.1016/S2468-1253(19)30327-9 / 31953079', 'file_path': 'https://pubmed.ncbi.nlm.nih.gov/31953079/',
        'search_phrase': None,
        'manual_quote': 'The primary endpoint was time to treatment failure; secondary endpoints were disease control rate, overall response rate, progression-free survival, overall survival, safety, and quality of life.',
        'manual_page': 'Abstract', 'evidence_level': 'B',
    },
    {
        'datum_id': 'E024', 'task_id': '任务1/2', 'trial_id': 'GEST', 'endpoint': '截图生存与ORR细项',
        'value_text': '三臂mOS/mPFS/1yOS/6mPFS/12mPFS与ORR/CR/DCR',
        'population': 'LAPC/MPC', 'arm_compare': 'Gem vs S-1 vs GS', 'source_type': 'user_screenshot',
        'doi/pmid': '', 'file_path': 'user_attachment_images', 'search_phrase': None,
        'manual_quote': '用户截图中GEST分表给出三臂生存与肿瘤控制细项，作为二次录入核验来源。', 'manual_page': '截图', 'evidence_level': 'C',
    },
    {
        'datum_id': 'E025', 'task_id': '任务1/2', 'trial_id': 'CONKO-007', 'endpoint': '截图补充项',
        'value_text': '截图含手术率/R0与局部复发、远处转移拆分值',
        'population': 'LAPC', 'arm_compare': 'CRT vs ChT', 'source_type': 'user_screenshot',
        'doi/pmid': '', 'file_path': 'user_attachment_images', 'search_phrase': None,
        'manual_quote': '用户截图含CONKO-007的手术率与复发/转移拆分字段，已与全文主结论交叉核验。', 'manual_page': '截图', 'evidence_level': 'C',
    },
]

AE_MATRIX_RAW = [
    {'临床试验': 'ACCORD 11', 'arm_compare': 'FOLFIRINOX vs Gem', 'AE术语': '发热性中性粒细胞减少', '分级': 'grade3/4', '干预组数值': '5.4', '对照组数值': '', '单位': '%', '是否关键AE': '是', 'datum_id': 'E002'},
    {'临床试验': 'MPACT', 'arm_compare': 'GnP vs Gem', 'AE术语': '中性粒细胞减少', '分级': 'grade≥3', '干预组数值': '38', '对照组数值': '27', '单位': '%', '是否关键AE': '是', 'datum_id': 'E003'},
    {'临床试验': 'MPACT', 'arm_compare': 'GnP vs Gem', 'AE术语': '疲劳', '分级': 'grade≥3', '干预组数值': '17', '对照组数值': '7', '单位': '%', '是否关键AE': '是', 'datum_id': 'E003'},
    {'临床试验': 'MPACT', 'arm_compare': 'GnP vs Gem', 'AE术语': '周围神经病变', '分级': 'grade≥3', '干预组数值': '17', '对照组数值': '1', '单位': '%', '是否关键AE': '是', 'datum_id': 'E003'},
    {'临床试验': 'MPACT', 'arm_compare': 'GnP vs Gem', 'AE术语': '发热性中性粒细胞减少', '分级': 'grade≥3', '干预组数值': '3', '对照组数值': '1', '单位': '%', '是否关键AE': '否', 'datum_id': 'E003'},
    {'临床试验': 'NEOPAN', 'arm_compare': 'mFOLFIRINOX vs Gem', 'AE术语': '严重不良事件', '分级': 'grade3/4', '干预组数值': '41', '对照组数值': '32', '单位': '%', '是否关键AE': '是', 'datum_id': 'E004'},
    {'临床试验': 'NAPOLI-3', 'arm_compare': 'NALIRIFOX vs GnP', 'AE术语': '治疗期不良事件', '分级': 'grade≥3', '干预组数值': '87', '对照组数值': '86', '单位': '%', '是否关键AE': '是', 'datum_id': 'E005'},
    {'临床试验': 'NAPOLI-3', 'arm_compare': 'NALIRIFOX vs GnP', 'AE术语': '治疗相关死亡', '分级': 'grade5', '干预组数值': '2', '对照组数值': '2', '单位': '%', '是否关键AE': '是', 'datum_id': 'E005'},
    {'临床试验': 'JCOG1611 (GENERATE)', 'arm_compare': 'mFOLFIRINOX/S-IROX vs GnP', 'AE术语': '厌食', '分级': 'grade3/4', '干预组数值': '23.3/27.5', '对照组数值': '5.0', '单位': '%', '是否关键AE': '是', 'datum_id': 'E006'},
    {'临床试验': 'AVENGER 500', 'arm_compare': 'mFFX+Devimistat vs FFX', 'AE术语': '总体不良事件', '分级': 'grade3', '干预组数值': '57.9', '对照组数值': '48.9', '单位': '%', '是否关键AE': '是', 'datum_id': 'E009'},
    {'临床试验': 'AVENGER 500', 'arm_compare': 'mFFX+Devimistat vs FFX', 'AE术语': '总体不良事件', '分级': 'grade4', '干预组数值': '22.0', '对照组数值': '28.5', '单位': '%', '是否关键AE': '是', 'datum_id': 'E009'},
    {'临床试验': 'NOTABLE', 'arm_compare': 'Gem+尼妥珠单抗 vs Gem', 'AE术语': '重度不良事件', '分级': 'grade4-5', '干预组数值': '0', '对照组数值': '0', '单位': '%', '是否关键AE': '是', 'datum_id': 'E010'},
    {'临床试验': 'HALO-301', 'arm_compare': 'GnP+PEGPH20 vs GnP', 'AE术语': '疲劳', '分级': 'grade≥3', '干预组数值': '16.0', '对照组数值': '9.6', '单位': '%', '是否关键AE': '是', 'datum_id': 'E011'},
    {'临床试验': 'PA.3', 'arm_compare': 'Gem+Erlotinib vs Gem', 'AE术语': '中性粒细胞减少', '分级': 'grade3/4', '干预组数值': '24', '对照组数值': '27', '单位': '%', '是否关键AE': '是', 'datum_id': 'E012'},
    {'临床试验': 'PA.3', 'arm_compare': 'Gem+Erlotinib vs Gem', 'AE术语': '血小板减少', '分级': 'grade3/4', '干预组数值': '10', '对照组数值': '11', '单位': '%', '是否关键AE': '否', 'datum_id': 'E012'},
    {'临床试验': 'KG4/2015', 'arm_compare': 'GemCap+GV1001 vs GemCap', 'AE术语': '总体不良事件', '分级': 'grade≥3', '干预组数值': '77.3', '对照组数值': '73.1', '单位': '%', '是否关键AE': '是', 'datum_id': 'E014'},
    {'临床试验': 'KG4/2015', 'arm_compare': 'GemCap+GV1001 vs GemCap', 'AE术语': '中性粒细胞减少', '分级': 'grade≥3', '干预组数值': '57.3', '对照组数值': '50.8', '单位': '%', '是否关键AE': '是', 'datum_id': 'E014'},
    {'临床试验': 'KG4/2015', 'arm_compare': 'GemCap+GV1001 vs GemCap', 'AE术语': '贫血', '分级': 'grade≥3', '干预组数值': '16.0', '对照组数值': '13.4', '单位': '%', '是否关键AE': '否', 'datum_id': 'E014'},
    {'临床试验': 'KG4/2015', 'arm_compare': 'GemCap+GV1001 vs GemCap', 'AE术语': '血小板减少', '分级': 'grade≥3', '干预组数值': '9.3', '对照组数值': '13.4', '单位': '%', '是否关键AE': '否', 'datum_id': 'E014'},
    {'临床试验': 'KG4/2015', 'arm_compare': 'GemCap+GV1001 vs GemCap', 'AE术语': '白细胞减少', '分级': 'grade≥3', '干预组数值': '12.0', '对照组数值': '9.0', '单位': '%', '是否关键AE': '否', 'datum_id': 'E014'},
    {'临床试验': 'CONKO-007', 'arm_compare': 'CRT vs ChT', 'AE术语': '贫血', '分级': 'grade3/4', '干预组数值': '5', '对照组数值': '2', '单位': '%', '是否关键AE': '否', 'datum_id': 'E015'},
    {'临床试验': 'CONKO-007', 'arm_compare': 'CRT vs ChT', 'AE术语': '白细胞减少', '分级': 'grade3/4', '干预组数值': '30', '对照组数值': '7', '单位': '%', '是否关键AE': '是', 'datum_id': 'E015'},
    {'临床试验': 'CONKO-007', 'arm_compare': 'CRT vs ChT', 'AE术语': '中性粒细胞减少', '分级': 'grade3/4', '干预组数值': '2', '对照组数值': '4', '单位': '%', '是否关键AE': '否', 'datum_id': 'E015'},
    {'临床试验': 'CONKO-007', 'arm_compare': 'CRT vs ChT', 'AE术语': '血小板减少', '分级': 'grade3/4', '干预组数值': '25', '对照组数值': '8', '单位': '%', '是否关键AE': '是', 'datum_id': 'E015'},
    {'临床试验': 'LAP-07', 'arm_compare': 'CRT vs ChT', 'AE术语': '总体毒性', '分级': 'grade3/4', '干预组数值': '未增加(除恶心)', '对照组数值': '基线对照', '单位': '文本', '是否关键AE': '是', 'datum_id': 'E016'},
    {'临床试验': 'GEST++(IPD)', 'arm_compare': 'GS vs Gem', 'AE术语': 'rash/腹泻/呕吐/中性粒细胞减少', '分级': 'grade≥3', '干预组数值': '更高', '对照组数值': '较低', '单位': '文本', '是否关键AE': '是', 'datum_id': 'E008'},
    {'临床试验': 'ACCORD11_QOL', 'arm_compare': 'FOLFIRINOX vs Gem', 'AE术语': '腹泻', '分级': '症状域变化', '干预组数值': '前2月显著升高', '对照组数值': '较低', '单位': '文本', '是否关键AE': '否', 'datum_id': 'E017'},
    {'临床试验': 'MPACT_NICE_CE', 'arm_compare': 'Nab-Pac+Gem vs Gem', 'AE术语': 'AE disutility', '分级': '模型参数', '干预组数值': '已纳入', '对照组数值': '已纳入', '单位': '文本', '是否关键AE': '否', 'datum_id': 'E020'},
    {'临床试验': 'LAPACT', 'arm_compare': 'single-arm', 'AE术语': '安全性', '分级': '研究次要终点', '干预组数值': '已评估', '对照组数值': '', '单位': '文本', '是否关键AE': '否', 'datum_id': 'E023'},
]

DOWNLOAD_RECORDS = [
    {'文献': 'TeloVac phase III', 'doi/pmid': '10.1016/S1470-2045(14)70236-0 / PMID:24954781', '来源': 'paperscraper', 'download_status': '失败', '本地路径': '/Users/alfred/Desktop/paper/downloads/new_papers/TeloVac_2014.pdf|/Users/alfred/Desktop/paper/downloads/new_papers/TeloVac_2014.xml', '是否已深度提取': '否', '备注': 'DOI存在但未抓取到可用全文'},
    {'文献': 'PANOVA-3 ASCO 2023', 'doi/pmid': '10.1200/jco.2023.41.4_suppl.tps770', '来源': 'paperscraper', 'download_status': '失败', '本地路径': '/Users/alfred/Desktop/paper/downloads/new_papers/PANOVA3_2023_ASCO.pdf|/Users/alfred/Desktop/paper/downloads/new_papers/PANOVA3_2023_ASCO.xml', '是否已深度提取': '否', '备注': '会议摘要，暂无可抓取全文'},
    {'文献': 'PANOVA-3 ASCO 2022', 'doi/pmid': '10.1200/jco.2022.40.4_suppl.tps629', '来源': 'paperscraper', 'download_status': '失败', '本地路径': '/Users/alfred/Desktop/paper/downloads/new_papers/PANOVA3_2022_ASCO.pdf|/Users/alfred/Desktop/paper/downloads/new_papers/PANOVA3_2022_ASCO.xml', '是否已深度提取': '否', '备注': '会议摘要，暂无可抓取全文'},
    {'文献': 'PANOVA phase2', 'doi/pmid': '10.1016/j.pan.2018.10.004', '来源': 'paperscraper', 'download_status': '失败', '本地路径': '/Users/alfred/Desktop/paper/downloads/new_papers/PANOVA_phase2_2019.pdf|/Users/alfred/Desktop/paper/downloads/new_papers/PANOVA_phase2_2019.xml', '是否已深度提取': '否', '备注': '未抓取到全文'},
    {'文献': 'GEST QOL ESMO Open', 'doi/pmid': '10.1136/esmoopen-2016-000151 / PMID:28761731', '来源': 'paperscraper', 'download_status': '成功(xml)', '本地路径': '/Users/alfred/Desktop/paper/downloads/new_papers/GEST_QOL_ESMO_2017.xml', '是否已深度提取': '是', '备注': '开放获取XML'},
    {'文献': 'KG4/2015 BJC', 'doi/pmid': '10.1038/s41416-023-02474-w / PMID:38278089', '来源': 'paperscraper', 'download_status': '成功(xml)', '本地路径': '/Users/alfred/Desktop/paper/downloads/new_papers/KG4_2015_2024.xml', '是否已深度提取': '是', '备注': '开放获取XML'},
    {'文献': 'ACCORD11 QoL companion', 'doi/pmid': '10.1200/JCO.2012.44.4869 / PMID:23213101', '来源': 'paper_hub download-doi', 'download_status': '失败', '本地路径': '/Users/alfred/Desktop/paper/downloads/new_papers/ACCORD11_QOL_2013.pdf|/Users/alfred/Desktop/paper/downloads/new_papers/ACCORD11_QOL_2013.xml', '是否已深度提取': '摘要级', '备注': '仅PubMed摘要可得'},
    {'文献': 'PA.3 QoL prognostic', 'doi/pmid': '10.1016/j.pan.2016.08.013 / PMID:27600995', '来源': 'paper_hub download-doi', 'download_status': '失败', '本地路径': '/Users/alfred/Desktop/paper/downloads/new_papers/PA3_QOL_2016.pdf|/Users/alfred/Desktop/paper/downloads/new_papers/PA3_QOL_2016.xml', '是否已深度提取': '摘要级', '备注': '仅PubMed摘要可得'},
    {'文献': 'MPACT real-world HRQoL', 'doi/pmid': '10.1007/s12029-016-9902-9 / PMID:28028766', '来源': 'paper_hub download-doi', 'download_status': '失败', '本地路径': '/Users/alfred/Desktop/paper/downloads/new_papers/MPACT_HRQoL_2017.pdf|/Users/alfred/Desktop/paper/downloads/new_papers/MPACT_HRQoL_2017.xml', '是否已深度提取': '摘要级', '备注': '仅PubMed摘要可得'},
    {'文献': 'NICE STA ERG (mPC)', 'doi/pmid': '10.1007/s40273-018-0646-1 / PMID:29600384', '来源': 'paper_hub download-doi', 'download_status': '成功(xml)', '本地路径': '/Users/alfred/Desktop/paper/downloads/new_papers/MPACT_NICE_CE_2018.xml', '是否已深度提取': '是', '备注': 'PMC BioC XML抓取成功'},
    {'文献': 'APICE', 'doi/pmid': '10.1080/14737167.2018.1464394 / PMID:29641931', '来源': 'paper_hub download-doi', 'download_status': '失败', '本地路径': '/Users/alfred/Desktop/paper/downloads/new_papers/APICE_2018.pdf|/Users/alfred/Desktop/paper/downloads/new_papers/APICE_2018.xml', '是否已深度提取': '摘要级', '备注': '摘要可用，全文抓取失败'},
    {'文献': 'FOLFIRINOX vs GEM-N CE', 'doi/pmid': '10.5301/tj.5000499 / PMID:27056335', '来源': 'paper_hub download-doi', 'download_status': '失败', '本地路径': '/Users/alfred/Desktop/paper/downloads/new_papers/FOLFIRINOX_vs_GNP_CE_2016.pdf|/Users/alfred/Desktop/paper/downloads/new_papers/FOLFIRINOX_vs_GNP_CE_2016.xml', '是否已深度提取': '摘要级', '备注': '摘要可用，全文抓取失败'},
    {'文献': 'LAPACT', 'doi/pmid': '10.1016/S2468-1253(19)30327-9 / PMID:31953079', '来源': 'paper_hub download-doi', 'download_status': '失败', '本地路径': '/Users/alfred/Desktop/paper/downloads/new_papers/LAPACT_2020.pdf|/Users/alfred/Desktop/paper/downloads/new_papers/LAPACT_2020.xml', '是否已深度提取': '摘要级', '备注': '摘要可用，全文抓取失败'},
]


# ----------------------------- extraction helpers -----------------------------

def normalize_text(text: str) -> str:
    return re.sub(r'\s+', ' ', text or '').strip()


def find_phrase_snippet(text: str, phrase: str, radius: int = 220) -> Optional[str]:
    if not text or not phrase:
        return None
    t = normalize_text(text)
    p = normalize_text(phrase)
    idx = t.lower().find(p.lower())
    if idx < 0:
        return None
    start = max(0, idx - radius)
    end = min(len(t), idx + len(p) + radius)
    return t[start:end]


def extract_with_pypdf(path: Path, phrase: str) -> Tuple[Optional[int], Optional[str]]:
    if not path.exists():
        return None, None
    try:
        reader = PdfReader(str(path))
    except Exception:
        return None, None
    for i, page in enumerate(reader.pages, start=1):
        txt = page.extract_text() or ''
        snippet = find_phrase_snippet(txt, phrase)
        if snippet:
            return i, snippet
    return None, None


def extract_with_pdfplumber(path: Path, phrase: str) -> Tuple[Optional[int], Optional[str]]:
    if not path.exists():
        return None, None
    try:
        with pdfplumber.open(str(path)) as pdf:
            for i, page in enumerate(pdf.pages, start=1):
                txt = page.extract_text() or ''
                snippet = find_phrase_snippet(txt, phrase)
                if snippet:
                    return i, snippet
    except Exception:
        return None, None
    return None, None


def extract_with_xml(path: Path, phrase: str) -> Tuple[Optional[str], Optional[str]]:
    if not path.exists():
        return None, None
    txt = path.read_text(encoding='utf-8', errors='ignore')
    snippet = find_phrase_snippet(txt, phrase)
    if not snippet:
        no_tags = re.sub(r'<[^>]+>', ' ', txt)
        snippet = find_phrase_snippet(no_tags, phrase)
    if snippet:
        return 'XML-ABSTRACT', snippet
    return None, None


def choose_quote(a: Optional[str], b: Optional[str], manual: Optional[str]) -> str:
    if manual:
        return manual
    if a and b:
        return a if len(a) >= len(b) else b
    return a or b or ''


def consistency(a: Optional[str], b: Optional[str]) -> str:
    if not a and not b:
        return '无自动抽取'
    if a and b:
        an = normalize_text(a).lower()
        bn = normalize_text(b).lower()
        if an == bn:
            return '一致'
        aset = set(an.split())
        bset = set(bn.split())
        union = len(aset | bset) or 1
        overlap = len(aset & bset) / union
        return '部分一致' if overlap >= 0.35 else '冲突'
    return '单通道命中'


# ----------------------------- dataframe builders -----------------------------

def extract_year(pub_text: str) -> int:
    m = re.search(r'(19\d{2}|20\d{2})', str(pub_text))
    return int(m.group(1)) if m else 9999


def sort_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out['_method'] = out['方式'].map(METHOD_ORDER).fillna(99)
    out['_year'] = out['发表时间'].map(extract_year)
    out['_layer'] = out['研究层级(core/extended)'].map(LAYER_ORDER).fillna(9)
    out = out.sort_values(['_method', '_year', '临床试验']).drop(columns=['_method', '_year', '_layer'])
    return out


def _task_dimension_id(task_name: str) -> str:
    if task_name == '任务1':
        return 'os_median'
    if task_name == '任务2':
        return 'orr'
    return 'ae_grade3plus'


def _infer_source_tier(evidence_grade: str) -> str:
    g = normalize_text(str(evidence_grade)).upper()
    if g.startswith('A'):
        return 'A'
    if g.startswith('B'):
        return 'B'
    if g.startswith('C'):
        return 'C'
    return 'C'


def attach_common(raw_rows: List[Dict[str, str]], task_name: str) -> pd.DataFrame:
    rows = []
    for r in raw_rows:
        trial = r['临床试验']
        meta = STUDIES[trial]
        dim_id = _task_dimension_id(task_name)
        base = {
            '方式': meta['方式'],
            '类别': meta['类别'],
            '临床试验': trial,
            '研究时间': meta['研究时间'],
            '随访时间/月': meta['随访时间/月'],
            '研究人群': meta['研究人群'],
            'n': meta['n'],
            '治疗方案': meta['治疗方案'],
            '发表时间': meta['发表时间'],
            '研究层级(core/extended)': meta['研究层级(core/extended)'],
            '是否新增文献': meta['是否新增文献'],
            'dimension_id': dim_id,
            'dimension_version': 'v1',
            'definition_source': f'dimensions_catalog.yaml::{dim_id}',
            'value_source': 'curated_clinical_evidence_bundle',
            'source_tier': _infer_source_tier(r.get('证据等级', 'C')),
            'institution_tier': '',
            'country_group': 'mixed',
        }
        base.update(r)
        rows.append(base)
    return sort_df(pd.DataFrame(rows))


def _dimension_from_endpoint(endpoint: str) -> str:
    t = normalize_text(endpoint).lower()
    if 'overall survival' in t or 'os' in t:
        return 'os_median'
    if 'progression' in t or 'pfs' in t:
        return 'pfs_median'
    if 'orr' in t or 'response' in t or 'dcr' in t:
        return 'orr'
    if 'qaly' in t or 'icer' in t or 'cost' in t:
        return 'qaly'
    if 'qol' in t or 'quality' in t or 'tudd' in t:
        return 'qol_score'
    if 'adverse' in t or 'ae' in t or 'ctcae' in t or 'death' in t:
        return 'ae_grade3plus'
    return 'custom_clinical_signal'


def _source_tier_from_source_type(source_type: str) -> str:
    st = normalize_text(source_type).lower()
    if st in {'fulltext_pdf', 'pmc_bioc_xml'}:
        return 'A'
    if st in {'pubmed_abstract', 'user_screenshot'}:
        return 'B'
    if st == 'missingness_audit':
        return 'C'
    return 'C'


def display_width(value: object) -> int:
    text = normalize_text(str(value or ''))
    if not text:
        return 0
    width = 0
    for ch in text:
        width += 2 if unicodedata.east_asian_width(ch) in {'W', 'F'} else 1
    return width


def style_sheet(ws) -> None:
    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill(fill_type='solid', fgColor=MCKINSEY_BLUE)
    header_font = Font(color='FFFFFF', bold=True)
    first_col_font = Font(color=MCKINSEY_BLUE, bold=True)
    base_font = Font(color='000000', bold=False)

    max_row = ws.max_row
    max_col = ws.max_column
    if max_row <= 0 or max_col <= 0:
        return

    for col_idx in range(1, max_col + 1):
        max_w = 0
        for row_idx in range(1, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            max_w = max(max_w, display_width(cell.value))

            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            elif col_idx == 1:
                cell.font = first_col_font
            else:
                cell.font = base_font

        width = min(60, max(10, max_w + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = 'A2'


def build_evidence_df() -> pd.DataFrame:
    out = []
    for spec in EVIDENCE_SPECS:
        source_type = spec['source_type']
        phrase = spec.get('search_phrase')
        file_path = spec.get('file_path') or ''
        manual_quote = spec.get('manual_quote')

        page_a = page_b = page_xml = None
        quote_a = quote_b = quote_xml = None

        if source_type == 'fulltext_pdf' and phrase and file_path:
            p = Path(file_path)
            page_a, quote_a = extract_with_pypdf(p, phrase)
            page_b, quote_b = extract_with_pdfplumber(p, phrase)
        elif source_type == 'pmc_bioc_xml' and phrase and file_path:
            page_xml, quote_xml = extract_with_xml(Path(file_path), phrase)

        chosen_quote = choose_quote(quote_a or quote_xml, quote_b, manual_quote)
        page_no = spec.get('manual_page') or page_a or page_b or page_xml or 'N/A'

        out.append({
            'datum_id': spec['datum_id'],
            'endpoint': spec['endpoint'],
            'value_text': spec['value_text'],
            'population': spec['population'],
            'arm_compare': spec['arm_compare'],
            'source_type': source_type,
            'doi/pmid': spec['doi/pmid'],
            'file_path': file_path,
            'page_no': page_no,
            'quote_original': chosen_quote,
            'evidence_level': spec['evidence_level'],
            'dimension_id': _dimension_from_endpoint(spec['endpoint']),
            'dimension_version': 'v1',
            'definition_source': f"dimensions_catalog.yaml::{_dimension_from_endpoint(spec['endpoint'])}",
            'value_source': source_type,
            'source_tier': _source_tier_from_source_type(source_type),
            'institution_tier': '',
            'country_group': 'mixed',
            'task_id': spec['task_id'],
            'trial_id': spec['trial_id'],
            'extractor_a_quote': quote_a or quote_xml or '',
            'extractor_b_quote': quote_b or '',
            'consistency': consistency(quote_a or quote_xml, quote_b),
            'adjudication_note': 'manual_quote优先；否则采用双提取器命中片段。',
        })

    cols = ['datum_id', 'endpoint', 'value_text', 'population', 'arm_compare', 'source_type', 'doi/pmid', 'file_path', 'page_no', 'quote_original', 'evidence_level', 'dimension_id', 'dimension_version', 'definition_source', 'value_source', 'source_tier', 'institution_tier', 'country_group', 'task_id', 'trial_id', 'extractor_a_quote', 'extractor_b_quote', 'consistency', 'adjudication_note']
    return pd.DataFrame(out)[cols]


def append_evidence_id(existing: str, new_id: str) -> str:
    parts = [p.strip() for p in str(existing).split(';') if p.strip()]
    if new_id not in parts:
        parts.append(new_id)
    return ';'.join(parts)


def add_missingness_evidence(task_df: pd.DataFrame, task_name: str, metric_cols: List[str], start_idx: int) -> Tuple[pd.DataFrame, List[Dict[str, str]], int]:
    rows = []
    out_df = task_df.copy()

    for i, row in out_df.iterrows():
        missing = [c for c in metric_cols if normalize_text(str(row.get(c, ''))) == '']
        if not missing:
            continue

        datum_id = f'M{start_idx:03d}'
        start_idx += 1

        rows.append({
            'datum_id': datum_id,
            'endpoint': f'{task_name}_缺失字段说明',
            'value_text': f"未报告字段: {', '.join(missing)}",
            'population': row['研究人群'],
            'arm_compare': row['治疗方案'],
            'source_type': 'missingness_audit',
            'doi/pmid': '',
            'file_path': '',
            'page_no': 'N/A',
            'quote_original': '在当前可得全文/摘要/补充材料中未检索到可核对数值，主表按规则留空。',
            'evidence_level': 'C',
            'dimension_id': row.get('dimension_id', _task_dimension_id(task_name)),
            'dimension_version': 'v1',
            'definition_source': f"dimensions_catalog.yaml::{row.get('dimension_id', _task_dimension_id(task_name))}",
            'value_source': 'missingness_audit',
            'source_tier': 'C',
            'institution_tier': '',
            'country_group': 'mixed',
            'task_id': task_name,
            'trial_id': row['临床试验'],
            'extractor_a_quote': '',
            'extractor_b_quote': '',
            'consistency': '人工裁决',
            'adjudication_note': '缺失值策略：主表留空并在底表登记。',
        })

        out_df.at[i, '证据ID'] = append_evidence_id(str(row['证据ID']), datum_id)

    return out_df, rows, start_idx


def build_total_sheet(task1: pd.DataFrame, task2: pd.DataFrame, task3: pd.DataFrame) -> pd.DataFrame:
    key_cols = ['方式', '类别', '临床试验', '研究时间', '随访时间/月', '研究人群', 'n', '治疗方案', '发表时间', '研究层级(core/extended)', '是否新增文献']
    t1 = task1[key_cols + ['mOS/月', 'mPFS/月', '1yOS/%', '12mPFS/%', 'DFS/%', '证据等级']].rename(columns={'证据等级': '任务1证据等级'})
    t2 = task2[['临床试验', 'ORR/%', 'DCR/%', 'R0切除率/%', '局部复发率/%', 'CBR/%', '证据等级']].rename(columns={'证据等级': '任务2证据等级'})
    t3 = task3[['临床试验', '≥3级AE/%', '治疗相关死亡/%', 'QOL', 'QALY/QALM', '证据等级']].rename(columns={'证据等级': '任务3证据等级'})
    out = t1.merge(t2, on='临床试验', how='left').merge(t3, on='临床试验', how='left')
    return sort_df(out)


def _pair(value: str) -> Tuple[Optional[float], Optional[float]]:
    nums = re.findall(r'\d+(?:\.\d+)?', str(value))
    if len(nums) < 2:
        return None, None
    return float(nums[0]), float(nums[1])


def build_surprises(task1: pd.DataFrame, task2: pd.DataFrame, task3: pd.DataFrame, evidence_df: pd.DataFrame) -> pd.DataFrame:
    merged = task1[['临床试验', '方式', '发表时间', 'mOS/月', 'mPFS/月', '证据ID']].merge(
        task2[['临床试验', 'ORR/%', '证据ID']].rename(columns={'证据ID': '证据ID_2'}), on='临床试验', how='left'
    ).merge(
        task3[['临床试验', '≥3级AE/%', 'QOL', '证据ID']].rename(columns={'证据ID': '证据ID_3'}), on='临床试验', how='left'
    )

    rows = []
    ev_map = evidence_df.set_index('datum_id').to_dict('index')

    for _, r in merged.iterrows():
        trial = r['临床试验']
        layer = STUDIES.get(trial, {}).get('研究层级(core/extended)', '')
        orr_t, orr_c = _pair(r.get('ORR/%', ''))
        os_t, os_c = _pair(r.get('mOS/月', ''))
        pfs_t, pfs_c = _pair(r.get('mPFS/月', ''))
        ae_t, ae_c = _pair(r.get('≥3级AE/%', ''))

        ids = ';'.join([x for x in [str(r.get('证据ID', '')), str(r.get('证据ID_2', '')), str(r.get('证据ID_3', ''))] if x])
        id_list = [x.strip() for x in ids.split(';') if x.strip()]
        first_quote = ''
        for i in id_list:
            if i in ev_map and ev_map[i].get('quote_original'):
                first_quote = ev_map[i]['quote_original']
                break

        if orr_t is not None and os_t is not None and orr_t > orr_c and os_t <= os_c:
            rows.append({'方式': r['方式'], '临床试验': trial, '发表时间': r['发表时间'], '研究层级(core/extended)': layer, '发现类型': 'ORR改善但OS不改善', '触发条件': f'ORR {r.get("ORR/%", "")} ; mOS {r.get("mOS/月", "")}', '证据ID': ids, '证据摘录': first_quote})

        if ae_t is not None and os_t is not None and ae_t > ae_c and os_t <= os_c:
            rows.append({'方式': r['方式'], '临床试验': trial, '发表时间': r['发表时间'], '研究层级(core/extended)': layer, '发现类型': '安全性更差且生存无获益', '触发条件': f'≥3级AE {r.get("≥3级AE/%", "")} ; mOS {r.get("mOS/月", "")}', '证据ID': ids, '证据摘录': first_quote})

        q = str(r.get('QOL', ''))
        if pfs_t is not None and pfs_t > pfs_c and any(k in q for k in ['恶化', '下降', 'worse', 'degradation']):
            rows.append({'方式': r['方式'], '临床试验': trial, '发表时间': r['发表时间'], '研究层级(core/extended)': layer, '发现类型': 'PFS改善但QOL变差', '触发条件': f'mPFS {r.get("mPFS/月", "")} ; QOL {q}', '证据ID': ids, '证据摘录': first_quote})

    rows.extend([
        {'方式': '放化疗', '临床试验': 'CONKO-007', '发表时间': '2025, JCO', '研究层级(core/extended)': 'core', '发现类型': '亚组结论与总体相反', '触发条件': '随机主比较OS无差异，但手术亚组mOS显著更长', '证据ID': 'E015;E025', '证据摘录': 'R0切除率与手术亚组生存在CRT方向更优，但ITT OS未显著改善。'},
        {'方式': '化疗', '临床试验': 'NEOPAN', '发表时间': '2025, JCO', '研究层级(core/extended)': 'core', '发现类型': 'PFS获益未转化为OS获益', '触发条件': 'mPFS提升但mOS近似持平', '证据ID': 'E004', '证据摘录': 'FOLFIRINOX显著提高PFS，但OS未显著提高。'},
        {'方式': '靶向治疗', '临床试验': 'HALO-301', '发表时间': '2020, JCO', '研究层级(core/extended)': 'core', '发现类型': 'ORR提升但OS/PFS均未提升', '触发条件': 'ORR 47 vs 36, mOS 11.2 vs 11.5, mPFS 7.1 vs 7.1', '证据ID': 'E011', '证据摘录': 'ORR改善并未转化为OS/PFS改善。'},
    ])

    out = pd.DataFrame(rows)
    if out.empty:
        return out
    return sort_df(out)


def build_ae_matrix(evidence_df: pd.DataFrame) -> pd.DataFrame:
    ev = evidence_df.set_index('datum_id').to_dict('index')
    rows = []
    for r in AE_MATRIX_RAW:
        trial = r['临床试验']
        meta = STUDIES[trial]
        datum = ev.get(r['datum_id'], {})
        rows.append({
            '方式': meta['方式'],
            '类别': meta['类别'],
            '临床试验': trial,
            '发表时间': meta['发表时间'],
            '研究层级(core/extended)': meta['研究层级(core/extended)'],
            'arm_compare': r['arm_compare'],
            'AE术语': r['AE术语'],
            '分级': r['分级'],
            '干预组数值': r['干预组数值'],
            '对照组数值': r['对照组数值'],
            '单位': r['单位'],
            '是否关键AE': r['是否关键AE'],
            'datum_id': r['datum_id'],
            'evidence_level': datum.get('evidence_level', ''),
            'page_no': datum.get('page_no', ''),
            'quote_original': datum.get('quote_original', ''),
        })
    return sort_df(pd.DataFrame(rows))


def build_error_sheet(task1: pd.DataFrame, task2: pd.DataFrame, evidence_df: pd.DataFrame) -> pd.DataFrame:
    checks = [
        {'临床试验': 'PANOVA-3', '任务': '任务1', '字段': '证据等级', '截图原值': 'A', '证据ID': 'E001', '说明': '当前仅会议摘要与截图，按B/C处理'},
        {'临床试验': 'CONKO-007', '任务': '任务1', '字段': '5yOS/%', '截图原值': '3.8 vs 10.1', '证据ID': 'E015;E025', '说明': '截图字段定义与全文随访定义不一致，主表留空并在备注说明'},
        {'临床试验': 'LAP-07', '任务': '任务2', '字段': '手术率/%', '截图原值': '6 vs 3', '证据ID': 'E016', '说明': '正文主证据聚焦局部进展与毒性，手术率字段未在主文稳定报告'},
        {'临床试验': 'GEST', '任务': '任务1', '字段': 'mOS/月', '截图原值': '8.8 vs 9.7 vs 10.1', '证据ID': 'E007;E024', '说明': '三臂值保留并标记含截图核验来源'},
        {'临床试验': 'TeloVac', '任务': '任务1', '字段': 'mOS/月', '截图原值': '7.9 vs 6.9 vs 8.4', '证据ID': 'E013', '说明': '全文不可得，保留截图值并降级证据'},
        {'临床试验': 'CONKO-007', '任务': '任务2', '字段': 'R0切除率/%', '截图原值': '18 vs 25；50 vs 69(手术人群)', '证据ID': 'E015;E025', '说明': '核验后拆分ITT与手术亚组两层'},
        {'临床试验': 'AVENGER 500', '任务': '任务1', '字段': 'mOS/月', '截图原值': '11.1 vs 11.73', '证据ID': 'E009', '说明': '与全文方向一致，标注为反直觉发现候选'},
    ]

    rows = []
    ev = evidence_df.set_index('datum_id').to_dict('index')

    for c in checks:
        trial = c['临床试验']
        field = c['字段']
        if c['任务'] == '任务1':
            current = task1.loc[task1['临床试验'] == trial, field].iloc[0] if field in task1.columns else ''
        else:
            current = task2.loc[task2['临床试验'] == trial, field].iloc[0] if field in task2.columns else ''

        ids = [x.strip() for x in c['证据ID'].split(';') if x.strip()]
        pages = []
        quotes = []
        for i in ids:
            if i in ev:
                pages.append(str(ev[i].get('page_no', '')))
                q = ev[i].get('quote_original', '')
                if q:
                    quotes.append(q)

        rows.append({
            '临床试验': trial,
            '任务': c['任务'],
            '字段': field,
            '截图原值': c['截图原值'],
            '核验后值': current,
            '状态': '一致' if str(c['截图原值']) == str(current) else '不一致/需解释',
            '证据ID': c['证据ID'],
            '原文页码': ','.join(pages[:2]),
            '原文摘录': ' || '.join(quotes[:2]),
            '说明': c['说明'],
        })

    return pd.DataFrame(rows)


def build_download_sheet() -> pd.DataFrame:
    df = pd.DataFrame(DOWNLOAD_RECORDS)

    def exists_flag(path_text: str) -> str:
        parts = [p.strip() for p in str(path_text).split('|') if p.strip()]
        if not parts:
            return 'N'
        return ','.join(['Y' if Path(p).exists() else 'N' for p in parts])

    df['本地存在性'] = df['本地路径'].apply(exists_flag)
    return df


def build_quality_sheets() -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    scoring_path = QUALITY_DIR / 'quality_scoring.csv'
    columns = [
        'uid', 'title', 'source', 'year', 'doi', 'pmid', 'pmcid', 'journal', 'discipline_profile',
        'credibility_score', 'credibility_tier', 'quality_gate', 'rejection_reason', 'journal_tier',
        'source_cred', 'journal_cred', 'citation_cred', 'design_cred', 'integrity_cred',
        'citation_age_years', 'citation_age_adjusted', 'cited_by_count', 'preprint_flag',
        'retracted_flag', 'institution_signal'
    ]
    if not scoring_path.exists():
        empty = pd.DataFrame(columns=columns)
        return empty.copy(), empty.copy(), empty.copy()

    quality = pd.read_csv(scoring_path)
    for c in columns:
        if c not in quality.columns:
            quality[c] = ''
    quality = quality[columns].copy()
    quality['credibility_score'] = pd.to_numeric(quality['credibility_score'], errors='coerce').fillna(0)
    quality = quality.sort_values(['quality_gate', 'credibility_score'], ascending=[True, False])

    rejected = quality[quality['quality_gate'] == 'reject'].copy()

    preprint_flag = quality['preprint_flag'].astype(str).str.lower().isin({'true', '1', 'yes'})
    preprint = quality[(quality['quality_gate'] == 'preprint_extended') | preprint_flag].copy()

    return quality, rejected, preprint


def build_field_dictionary(sheet_frames: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    field_to_sheets: Dict[str, set] = {}
    field_examples: Dict[str, str] = {}

    for sheet_name, df in sheet_frames.items():
        for col in df.columns:
            field_to_sheets.setdefault(col, set()).add(sheet_name)
            if col in field_examples:
                continue
            series = df[col].astype(str).map(normalize_text)
            series = series[series != '']
            if not series.empty:
                field_examples[col] = series.iloc[0][:160]

    rows = []
    ordered_fields = sorted(
        field_to_sheets.keys(),
        key=lambda x: (
            FIELD_PRIORITY_RANK.get(x, 10_000),
            0 if x in FIELD_GLOSSARY else 1,
            x,
        ),
    )
    for field in ordered_fields:
        rows.append(
            {
                '字段名': field,
                '字段含义': FIELD_GLOSSARY.get(field, '当前版本未配置自动解释，需在FIELD_GLOSSARY补充。'),
                '出现Sheet': '、'.join(sorted(field_to_sheets[field])),
                '示例值': field_examples.get(field, ''),
            }
        )

    return pd.DataFrame(rows, columns=['字段名', '字段含义', '出现Sheet', '示例值'])


def validate(task1: pd.DataFrame, task2: pd.DataFrame, task3: pd.DataFrame, evidence_df: pd.DataFrame, error_df: pd.DataFrame, field_dict_df: pd.DataFrame, sheets_written: List[str]) -> None:
    expected_sheets = [
        '字段词典', '任务1_生存终点', '任务2_肿瘤控制', '任务3_生活质量', '总表_集中', '任务2_按方式', '任务3_按方式',
        '任务3_AE矩阵', '底表_原文证据', '错误校验_截图', '新增论文_下载', '质量评分_总表',
        '剔除文献_不合格', '预印本_扩展待核', '反直觉发现', '任务2_化疗分层', '任务3_化疗分层'
    ]
    assert sheets_written == expected_sheets, f'sheet列表不匹配: {sheets_written}'

    ev_ids = set(evidence_df['datum_id'].astype(str))
    check_cols = TASK3_METRICS
    for _, row in task3.iterrows():
        ids = [x.strip() for x in str(row['证据ID']).split(';') if x.strip()]
        for c in check_cols:
            if normalize_text(str(row.get(c, ''))):
                assert ids, f'任务3非空字段缺证据ID: {row["临床试验"]} {c}'
                assert any(i in ev_ids for i in ids), f'任务3证据ID未落底表: {row["临床试验"]} {c}'

    for df, name, metrics in [(task1, '任务1', TASK1_METRICS), (task2, '任务2', TASK2_METRICS), (task3, '任务3', TASK3_METRICS)]:
        for _, row in df.iterrows():
            has_missing = any(normalize_text(str(row.get(c, ''))) == '' for c in metrics)
            if has_missing:
                ids = [x.strip() for x in str(row['证据ID']).split(';') if x.strip()]
                assert any(i.startswith('M') for i in ids), f'{name}空值未登记缺失证据: {row["临床试验"]}'

    assert 'CBR/%' in task2.columns, '任务2缺少CBR'
    assert 'CBR/%' not in task1.columns, '任务1不应包含CBR'
    assert 'CBR/%' not in task3.columns, '任务3不应包含CBR'

    assert len(error_df) >= 6, '错误校验sheet行数异常'
    assert not field_dict_df.empty, '字段词典为空'
    assert {'字段名', '字段含义', '出现Sheet', '示例值'} <= set(field_dict_df.columns), '字段词典结构异常'


def main() -> None:
    task1 = attach_common(TASK1_ROWS_RAW, '任务1')
    task2 = attach_common(TASK2_ROWS_RAW, '任务2')
    task3 = attach_common(TASK3_ROWS_RAW, '任务3')

    evidence_df = build_evidence_df()

    missing_idx = 1
    task1, miss1, missing_idx = add_missingness_evidence(task1, '任务1', TASK1_METRICS, missing_idx)
    task2, miss2, missing_idx = add_missingness_evidence(task2, '任务2', TASK2_METRICS, missing_idx)
    task3, miss3, missing_idx = add_missingness_evidence(task3, '任务3', TASK3_METRICS, missing_idx)

    if miss1 or miss2 or miss3:
        evidence_df = pd.concat([evidence_df, pd.DataFrame(miss1 + miss2 + miss3)], ignore_index=True)

    # keep required ordering in evidence sheet
    evidence_df = evidence_df[['datum_id', 'endpoint', 'value_text', 'population', 'arm_compare', 'source_type', 'doi/pmid', 'file_path', 'page_no', 'quote_original', 'evidence_level', 'dimension_id', 'dimension_version', 'definition_source', 'value_source', 'source_tier', 'institution_tier', 'country_group', 'task_id', 'trial_id', 'extractor_a_quote', 'extractor_b_quote', 'consistency', 'adjudication_note']]

    task1 = sort_df(task1)
    task2 = sort_df(task2)
    task3 = sort_df(task3)

    total_sheet = build_total_sheet(task1, task2, task3)
    task2_by_method = sort_df(task2.copy())
    task3_by_method = sort_df(task3.copy())
    ae_matrix = build_ae_matrix(evidence_df)
    error_sheet = build_error_sheet(task1, task2, evidence_df)
    downloads_sheet = build_download_sheet()
    quality_sheet, rejected_sheet, preprint_sheet = build_quality_sheets()
    surprise_sheet = build_surprises(task1, task2, task3, evidence_df)

    task2_chemo = task2[task2['方式'] == '化疗'].copy()
    task2_chemo['_layer'] = task2_chemo['研究层级(core/extended)'].map(LAYER_ORDER).fillna(9)
    task2_chemo['_year'] = task2_chemo['发表时间'].map(extract_year)
    task2_chemo = task2_chemo.sort_values(['_layer', '_year', '临床试验']).drop(columns=['_layer', '_year'])

    task3_chemo = task3[task3['方式'] == '化疗'].copy()
    task3_chemo['_layer'] = task3_chemo['研究层级(core/extended)'].map(LAYER_ORDER).fillna(9)
    task3_chemo['_year'] = task3_chemo['发表时间'].map(extract_year)
    task3_chemo = task3_chemo.sort_values(['_layer', '_year', '临床试验']).drop(columns=['_layer', '_year'])

    sheet_frames = {
        '任务1_生存终点': task1,
        '任务2_肿瘤控制': task2,
        '任务3_生活质量': task3,
        '总表_集中': total_sheet,
        '任务2_按方式': task2_by_method,
        '任务3_按方式': task3_by_method,
        '任务3_AE矩阵': ae_matrix,
        '底表_原文证据': evidence_df,
        '错误校验_截图': error_sheet,
        '新增论文_下载': downloads_sheet,
        '质量评分_总表': quality_sheet,
        '剔除文献_不合格': rejected_sheet,
        '预印本_扩展待核': preprint_sheet,
        '反直觉发现': surprise_sheet,
        '任务2_化疗分层': task2_chemo,
        '任务3_化疗分层': task3_chemo,
    }
    field_dict = build_field_dictionary(sheet_frames)

    sheet_order = [
        '字段词典', '任务1_生存终点', '任务2_肿瘤控制', '任务3_生活质量', '总表_集中', '任务2_按方式', '任务3_按方式',
        '任务3_AE矩阵', '底表_原文证据', '错误校验_截图', '新增论文_下载', '质量评分_总表',
        '剔除文献_不合格', '预印本_扩展待核', '反直觉发现', '任务2_化疗分层', '任务3_化疗分层'
    ]

    with pd.ExcelWriter(OUT_XLSX, engine='openpyxl') as writer:
        field_dict.to_excel(writer, sheet_name='字段词典', index=False)
        task1.to_excel(writer, sheet_name='任务1_生存终点', index=False)
        task2.to_excel(writer, sheet_name='任务2_肿瘤控制', index=False)
        task3.to_excel(writer, sheet_name='任务3_生活质量', index=False)
        total_sheet.to_excel(writer, sheet_name='总表_集中', index=False)
        task2_by_method.to_excel(writer, sheet_name='任务2_按方式', index=False)
        task3_by_method.to_excel(writer, sheet_name='任务3_按方式', index=False)
        ae_matrix.to_excel(writer, sheet_name='任务3_AE矩阵', index=False)
        evidence_df.to_excel(writer, sheet_name='底表_原文证据', index=False)
        error_sheet.to_excel(writer, sheet_name='错误校验_截图', index=False)
        downloads_sheet.to_excel(writer, sheet_name='新增论文_下载', index=False)
        quality_sheet.to_excel(writer, sheet_name='质量评分_总表', index=False)
        rejected_sheet.to_excel(writer, sheet_name='剔除文献_不合格', index=False)
        preprint_sheet.to_excel(writer, sheet_name='预印本_扩展待核', index=False)
        surprise_sheet.to_excel(writer, sheet_name='反直觉发现', index=False)
        task2_chemo.to_excel(writer, sheet_name='任务2_化疗分层', index=False)
        task3_chemo.to_excel(writer, sheet_name='任务3_化疗分层', index=False)

        for name in sheet_order:
            style_sheet(writer.book[name])

    validate(task1, task2, task3, evidence_df, error_sheet, field_dict, sheet_order)

    print(f'Wrote: {OUT_XLSX}')
    print(f'Task1 rows={len(task1)} Task2 rows={len(task2)} Task3 rows={len(task3)}')
    print(f'AE matrix rows={len(ae_matrix)} Evidence rows={len(evidence_df)}')


if __name__ == '__main__':
    main()
